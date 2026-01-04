import tkinter as tk
from tkinter import scrolledtext, messagebox, ttk
import tkinter.font as tkfont
from openpyxl import load_workbook, Workbook
import os
import re

XLSX_FILE = "rules.xlsx"

# 외래 기록 출력 마지막 고정 문구(요청 반영)
FINAL_OUTPATIENT_NOTE = "환자 P/E 및 검사 결과 확인하였으며, 검사 결과 설명하고, 경과 악화 시 병원 내원할 것 설명함"


# ===== 전역 글자 크기 제어 =====
CURRENT_FONT_SIZE = 10

def _apply_named_font_sizes(size: int):
    """Tk 기본 named font들을 한 번에 갱신"""
    for name in ["TkDefaultFont", "TkTextFont", "TkFixedFont", "TkMenuFont", "TkHeadingFont", "TkCaptionFont", "TkSmallCaptionFont", "TkIconFont", "TkTooltipFont"]:
        try:
            f = tkfont.nametofont(name)
            f.configure(size=size)
        except Exception:
            pass

def _update_widget_fonts(widget: tk.Widget, size: int):
    """이미 만들어진 위젯 중, tuple font를 쓰는 것들을 재귀적으로 크기만 조정"""
    try:
        f = widget.cget("font")
        if isinstance(f, (tuple, list)) and len(f) >= 2 and isinstance(f[1], int):
            fam = f[0]
            rest = f[2:]
            widget.configure(font=(fam, size, *rest))
    except Exception:
        pass

    for child in widget.winfo_children():
        _update_widget_fonts(child, size)

def set_global_font_size(root: tk.Tk, size: int):
    global CURRENT_FONT_SIZE
    try:
        size = int(size)
    except Exception:
        return
    size = max(8, min(20, size))
    CURRENT_FONT_SIZE = size

    _apply_named_font_sizes(size)
    _update_widget_fonts(root, size)

    for w in root.winfo_children():
        if isinstance(w, tk.Toplevel):
            _update_widget_fonts(w, size)



# ===== 출력에서 불필요한 출처/파일명 라인 제거(표시 단계 전처리) =====
def strip_sources(text: str) -> str:
    if not text:
        return text
    patterns = [
        r".*ILD\s*진료지침.*\n?",                 # 예: 'ILD 진료지침 ...'
        r".*ILD\+.*진료지침.*\(2023.*\).*?\n?", # 예: 'ILD+진료지침+개정+(2023년)'
    ]
    for p in patterns:
        text = re.sub(p, "", text, flags=re.IGNORECASE)
    text = re.sub(r"\n{3,}", "\n\n", text).strip()
    return text



# ================= Lung TNM v8 Calculator (NSCLC) ================= #

def _tnm_stage_group(T: str, N: str, M: str) -> str:
    """AJCC 8th (NSCLC) stage grouping for common T categories used in this helper."""
    # Metastatic first
    if M == "M1c":
        return "Stage IVB"
    if M in ("M1a", "M1b"):
        return "Stage IVA"
    if M != "M0":
        return "Check TNM"

    # Handle unknown T
    if T == "TX" or not T:
        # Return a practical range based on N (similar to CanStaging feel)
        if N == "N0":
            return "Stage I–III (T에 따라)"
        if N == "N1":
            return "Stage IIB–IIIA"
        if N == "N2":
            return "Stage IIIA–IIIB"
        if N == "N3":
            return "Stage IIIB–IIIC"
        return "Stage I–III (T에 따라)"

    # Normalize T1mi as IA1 with N0
    # Exact stage mapping (simplified but aligned with earlier tool)
    if N == "N0":
        if T == "Tis":
            return "Stage 0"
        if T in ("T1mi", "T1a"):
            return "Stage IA1"
        if T == "T1b":
            return "Stage IA2"
        if T == "T1c":
            return "Stage IA3"
        if T == "T2a":
            return "Stage IB"
        if T == "T2b":
            return "Stage IIA"
        if T == "T3":
            return "Stage IIB"
        if T == "T4":
            return "Stage IIIA"
    if N == "N1":
        if T in ("T1mi", "T1a", "T1b", "T1c", "T2a", "T2b"):
            return "Stage IIB"
        if T == "T3":
            return "Stage IIIA"
        if T == "T4":
            return "Stage IIIA"
    if N == "N2":
        if T in ("T1mi", "T1a", "T1b", "T1c", "T2a", "T2b"):
            return "Stage IIIA"
        if T == "T3":
            return "Stage IIIB"
        if T == "T4":
            return "Stage IIIB"
    if N == "N3":
        if T in ("T1mi", "T1a", "T1b", "T1c", "T2a"):
            return "Stage IIIB"
        if T in ("T2b", "T3", "T4"):
            return "Stage IIIC"
    return "Check TNM"


class LungTNMCalculator(tk.Toplevel):
    """
    CanStaging 느낌의 '체크 선택 기반' TNM 계산기.
    - Tab1: T feature 체크 + size 입력 → Computed T
    - Tab2: N/M feature 체크 → Computed N/M + Stage
    """
    def __init__(self, master, open_rule_callback=None):
        super().__init__(master)
        self.title("Lung TNM v8 계산기 (NSCLC)")
        self.geometry("900x650")

        self.open_rule_callback = open_rule_callback

        self.current_T = "TX"

        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True)

        self.tab_t = ttk.Frame(self.notebook)
        self.tab_nm = ttk.Frame(self.notebook)

        self.notebook.add(self.tab_t, text="1) T 결정")
        self.notebook.add(self.tab_nm, text="2) N/M & Stage")

        self._build_tab_t()
        self._build_tab_nm()

    # ---------- Tab1: T ----------
    def _build_tab_t(self):
        f = self.tab_t
        f.columnconfigure(0, weight=1)
        f.columnconfigure(1, weight=0)

        ttk.Label(f, text="T 결정 (체크 선택 기반 자동 계산)", font=("맑은 고딕", 11, "bold")).grid(
            row=0, column=0, sticky="w", pady=(10, 8)
        )

        self.t_mia = tk.BooleanVar(value=False)
        self.t_main_bronchus = tk.BooleanVar(value=False)
        self.t_visceral_pleura = tk.BooleanVar(value=False)
        self.t_atelectasis = tk.BooleanVar(value=False)
        self.t_chest_wall = tk.BooleanVar(value=False)
        self.t_same_lobe = tk.BooleanVar(value=False)
        self.t_diff_lobe = tk.BooleanVar(value=False)
        self.t_critical_organs = tk.BooleanVar(value=False)

        self.t_size_str = tk.StringVar(value="")  # keep as string to avoid 0.0 default

        # --- Auto T 계산: 체크/크기 입력이 바뀌면 즉시 재계산 ---
        def _auto_compute_T(*_args):
            try:
                self.compute_T()
            except Exception:
                # 입력 중 일시적으로 파싱이 안 되는 경우 등은 무시
                pass

        for _v in [self.t_mia, self.t_main_bronchus, self.t_visceral_pleura, self.t_atelectasis,
                   self.t_chest_wall, self.t_same_lobe, self.t_diff_lobe, self.t_critical_organs]:
            _v.trace_add('write', _auto_compute_T)
        self.t_size_str.trace_add('write', _auto_compute_T)

        r = 1
        ttk.Checkbutton(f, text="Minimally invasive adenocarcinoma (MIA)", variable=self.t_mia).grid(row=r, column=0, sticky="w"); r += 1

        size_frame = ttk.Frame(f)
        size_frame.grid(row=r, column=0, sticky="w", pady=(2, 6))
        ttk.Label(size_frame, text="Invasive tumour size (cm): ").pack(side="left")
        ttk.Entry(size_frame, textvariable=self.t_size_str, width=10).pack(side="left")
        ttk.Label(size_frame, text="(예: 2.4)").pack(side="left", padx=(8, 0))
        r += 1

        ttk.Checkbutton(f, text="Main bronchus involvement", variable=self.t_main_bronchus).grid(row=r, column=0, sticky="w"); r += 1
        ttk.Checkbutton(f, text="Visceral pleura invasion", variable=self.t_visceral_pleura).grid(row=r, column=0, sticky="w"); r += 1
        ttk.Checkbutton(f, text="Atelectasis / obstructive pneumonitis", variable=self.t_atelectasis).grid(row=r, column=0, sticky="w"); r += 1
        ttk.Checkbutton(f, text="Chest wall / parietal pleura / pericardium / phrenic nerve", variable=self.t_chest_wall).grid(row=r, column=0, sticky="w"); r += 1
        ttk.Checkbutton(f, text="Separate tumour nodule in same lobe", variable=self.t_same_lobe).grid(row=r, column=0, sticky="w"); r += 1
        ttk.Checkbutton(f, text="Separate tumour nodule in different ipsilateral lobe", variable=self.t_diff_lobe).grid(row=r, column=0, sticky="w"); r += 1
        ttk.Checkbutton(
            f,
            text="Invasion: diaphragm / mediastinum / heart / great vessels / RLN / vertebra / esophagus",
            variable=self.t_critical_organs,
        ).grid(row=r, column=0, sticky="w"); r += 1

        btn_frame = ttk.Frame(f)
        btn_frame.grid(row=r, column=0, sticky="w", pady=(12, 6))
        ttk.Button(btn_frame, text="T 계산", command=self.compute_T).pack(side="left")
        ttk.Button(btn_frame, text="→ 다음(N/M)", command=lambda: self.notebook.select(self.tab_nm)).pack(side="left", padx=(10, 0))

        self.lbl_t = ttk.Label(f, text="Computed T: TX", foreground="blue", font=("맑은 고딕", 10, "bold"))
        self.lbl_t.grid(row=r, column=0, sticky="w", padx=(160, 0))
        r += 1

        self.lbl_t_note = ttk.Label(
            f,
            text="* 체크된 항목 중 가장 높은 T 기준으로 자동 결정(T4→T3→T2→T1).",
            foreground="#444444",
        )
        self.lbl_t_note.grid(row=r, column=0, sticky="w", pady=(8, 0))
        r += 1

    def _parse_size(self):
        s = (self.t_size_str.get() or "").strip()
        if not s:
            return None
        try:
            return float(s)
        except ValueError:
            return None

    def compute_T(self):
        size = self._parse_size()

        # Priority: T4 features
        if self.t_critical_organs.get() or self.t_diff_lobe.get():
            T = "T4"
        # T3 features
        elif self.t_chest_wall.get() or self.t_same_lobe.get():
            T = "T3"
        # T2 features
        elif self.t_main_bronchus.get() or self.t_visceral_pleura.get() or self.t_atelectasis.get() or (size is not None and size > 3):
            # size threshold for T2a/T2b: 4cm cut in this simplified logic
            if size is not None and size > 4:
                T = "T2b"
            else:
                T = "T2a"
        else:
            # Size-based T1
            if size is None:
                T = "TX"
            else:
                if self.t_mia.get() and size <= 0.5:
                    T = "T1mi"
                elif size <= 1:
                    T = "T1a"
                elif size <= 2:
                    T = "T1b"
                elif size <= 3:
                    T = "T1c"
                else:
                    T = "TX"

        self.current_T = T
        self.lbl_t.config(text=f"Computed T: {T}")
        self.lbl_t_nm.config(text=f"Computed T (from Tab1): {T}")
        self.update_stage()

    # ---------- Tab2: N/M & Stage ----------
    def _build_tab_nm(self):
        f = self.tab_nm
        f.columnconfigure(0, weight=1)

        self.lbl_t_nm = ttk.Label(f, text="Computed T (from Tab1): TX", font=("맑은 고딕", 10, "bold"))
        self.lbl_t_nm.grid(row=0, column=0, sticky="w", pady=(10, 6))

        # N
        ttk.Label(f, text="N (Regional lymph nodes)", font=("맑은 고딕", 10, "bold")).grid(row=1, column=0, sticky="w", pady=(6, 2))

        self.n_peri = tk.BooleanVar(value=False)
        self.n_medi = tk.BooleanVar(value=False)
        self.n_contra = tk.BooleanVar(value=False)
        self.n_supra = tk.BooleanVar(value=False)

        ttk.Checkbutton(f, text="Ipsilateral peribronchial / hilar", variable=self.n_peri, command=self.update_stage).grid(row=2, column=0, sticky="w")
        ttk.Checkbutton(f, text="Ipsilateral mediastinal / subcarinal", variable=self.n_medi, command=self.update_stage).grid(row=3, column=0, sticky="w")
        ttk.Checkbutton(f, text="Contralateral mediastinal / hilar", variable=self.n_contra, command=self.update_stage).grid(row=4, column=0, sticky="w")
        ttk.Checkbutton(f, text="Supraclavicular nodes", variable=self.n_supra, command=self.update_stage).grid(row=5, column=0, sticky="w")

        self.lbl_n = ttk.Label(f, text="Computed N: N0", foreground="blue", font=("맑은 고딕", 10, "bold"))
        self.lbl_n.grid(row=6, column=0, sticky="w", pady=(4, 0))
        self.lbl_n_desc = ttk.Label(f, text="N0: 림프절 전이 없음", foreground="#444444")
        self.lbl_n_desc.grid(row=7, column=0, sticky="w", pady=(0, 8))

        # M
        ttk.Label(f, text="M (Distant metastases)", font=("맑은 고딕", 10, "bold")).grid(row=8, column=0, sticky="w", pady=(6, 2))

        self.m_contra_lung = tk.BooleanVar(value=False)
        self.m_pleural = tk.BooleanVar(value=False)
        self.m_single = tk.BooleanVar(value=False)
        self.m_multi = tk.BooleanVar(value=False)

        ttk.Checkbutton(f, text="Contralateral lung nodule", variable=self.m_contra_lung, command=self.update_stage).grid(row=9, column=0, sticky="w")
        ttk.Checkbutton(f, text="Pleural/pericardial nodule or malignant effusion", variable=self.m_pleural, command=self.update_stage).grid(row=10, column=0, sticky="w")
        ttk.Checkbutton(f, text="Single extrathoracic metastasis (single lesion)", variable=self.m_single, command=self.update_stage).grid(row=11, column=0, sticky="w")
        ttk.Checkbutton(f, text="Multiple extrathoracic metastases", variable=self.m_multi, command=self.update_stage).grid(row=12, column=0, sticky="w")

        self.lbl_m = ttk.Label(f, text="Computed M: M0", foreground="blue", font=("맑은 고딕", 10, "bold"))
        self.lbl_m.grid(row=13, column=0, sticky="w", pady=(4, 0))
        self.lbl_m_desc = ttk.Label(f, text="M0: 원격 전이 없음", foreground="#444444")
        self.lbl_m_desc.grid(row=14, column=0, sticky="w", pady=(0, 10))

        # Stage
        ttk.Label(f, text="Stage 결과:", font=("맑은 고딕", 11, "bold")).grid(row=15, column=0, sticky="w", pady=(6, 2))
        self.lbl_stage = ttk.Label(f, text="Stage I–III (T에 따라)", foreground="darkred", font=("맑은 고딕", 11, "bold"))
        self.lbl_stage.grid(row=16, column=0, sticky="w")

        # Stage에 따른 규칙 열기 (메인 창과 연동)
        self.btn_open_stage_rule = ttk.Button(
            f,
            text="Stage 결과에 따른 항목 열기",
            command=self.open_stage_rule
        )
        self.btn_open_stage_rule.grid(row=16, column=0, sticky="e", padx=(0, 18))

        # Buttons
        btn = ttk.Frame(f)
        btn.grid(row=17, column=0, sticky="e", pady=(14, 10))
        ttk.Button(btn, text="← 이전(T)", command=lambda: self.notebook.select(self.tab_t)).pack(side="left")
        ttk.Button(btn, text="닫기", command=self.destroy).pack(side="left", padx=(10, 0))

        # Initialize
        self.update_stage()

    def compute_N(self) -> str:
        if self.n_contra.get() or self.n_supra.get():
            return "N3"
        if self.n_medi.get():
            return "N2"
        if self.n_peri.get():
            return "N1"
        return "N0"

    def compute_M(self) -> str:
        if self.m_multi.get():
            return "M1c"
        if self.m_single.get():
            return "M1b"
        if self.m_contra_lung.get() or self.m_pleural.get():
            return "M1a"
        return "M0"

    def update_stage(self):
        N = self.compute_N()
        M = self.compute_M()
        T = self.current_T

        self.lbl_n.config(text=f"Computed N: {N}")
        self.lbl_m.config(text=f"Computed M: {M}")

        n_desc = {
            "N0": "N0: 림프절 전이 없음",
            "N1": "N1: 같은 쪽 폐문/기관지 주위/폐내 림프절",
            "N2": "N2: 같은 쪽 종격동 또는 carinal 림프절",
            "N3": "N3: 반대쪽 종격동/폐문 또는 쇄골상 림프절",
        }
        m_desc = {
            "M0": "M0: 원격 전이 없음",
            "M1a": "M1a: 반대쪽 폐 결절 또는 흉막/심낭 병변/삼출",
            "M1b": "M1b: 단일 장기 단일 전이 병소",
            "M1c": "M1c: 다발 원격 전이",
        }
        self.lbl_n_desc.config(text=n_desc.get(N, ""))
        self.lbl_m_desc.config(text=m_desc.get(M, ""))

        stage = _tnm_stage_group(T, N, M)
        self.lbl_stage.config(text=stage)



    def open_stage_rule(self):
        """현재 Stage 결과에 해당하는 규칙을 메인 창에서 열고(호흡기도우미), 메인 창으로 포커스 이동"""
        if not callable(self.open_rule_callback):
            messagebox.showinfo("안내", "메인 창과 연동 콜백이 설정되어 있지 않습니다.")
            return

        stage = (self.lbl_stage.cget("text") or "").strip()

        def _return_to_main():
            # 메인(호흡기도우미) 창을 앞으로
            try:
                self.master.deiconify()
            except Exception:
                pass
            try:
                self.master.lift()
                self.master.focus_force()
                # Windows에서 포커스가 안 올라오는 경우가 있어 topmost 토글
                self.master.attributes("-topmost", True)
                self.master.after(50, lambda: self.master.attributes("-topmost", False))
            except Exception:
                pass
            # TNM 창은 닫아 '돌아가기' 느낌을 줌 (원치 않으면 이 줄을 주석 처리)
            try:
                self.destroy()
            except Exception:
                pass

        # TX 등으로 range가 나오는 경우
        if stage.startswith("Stage I–III") or "–" in stage:
            messagebox.showinfo(
                "안내",
                f"현재 결과가 범위로 표시됩니다: {stage}\n(T를 확정하면 정확한 stage로 안내됩니다.)"
            )
            # 그래도 사용자 편의를 위해 범위 규칙을 열도록 시도
            self.open_rule_callback(stage)
            _return_to_main()
            return

        # 단일 stage인 경우
        self.open_rule_callback(stage)
        _return_to_main()
        return


def open_lung_tnm(master):
    try:
        LungTNMCalculator(master, open_rule_callback=open_rule_by_stage)
    except Exception as e:
        messagebox.showerror("Lung TNM 오류", f"TNM 창 실행 중 오류가 발생했습니다:\\n{e}")


# ================= 데이터 로딩 / 저장 ================= #

def ensure_rules_file(path=XLSX_FILE):
    """rules.xlsx가 없으면 헤더만 있는 새 파일 생성"""
    if os.path.exists(path):
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "rules"
    headers = ["category", "name", "keywords", "advice"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    wb.save(path)
    print(f">>> 새 {path} 생성 (헤더만 포함)")


def load_rules_from_xlsx(path=XLSX_FILE):
    """rules.xlsx에서 규칙 읽기"""
    ensure_rules_file(path)

    try:
        wb = load_workbook(path)
        ws = wb.active

        # 헤더 위치 확인
        header_map = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val:
                header_map[str(val).strip()] = col

        required = ["category", "name", "keywords", "advice"]
        for key in required:
            if key not in header_map:
                raise ValueError(f"엑셀에 '{key}' 열이 없습니다. (1행 헤더 확인)")

        rules_local = []
        for row in range(2, ws.max_row + 1):
            category = ws.cell(row=row, column=header_map["category"]).value
            name = ws.cell(row=row, column=header_map["name"]).value
            keywords_raw = ws.cell(row=row, column=header_map["keywords"]).value
            advice_raw = ws.cell(row=row, column=header_map["advice"]).value

            if not category and not name:
                continue

            category = str(category).strip() if category else ""
            name = str(name).strip() if name else ""
            keywords_raw = str(keywords_raw).strip() if keywords_raw else ""
            advice_raw = str(advice_raw) if advice_raw else ""

            if not category or not name:
                continue

            keywords = [
                kw.strip()
                for kw in keywords_raw.split(",")
                if kw.strip()
            ]

            rules_local.append(
                {
                    "category": category,
                    "name": name,
                    "keywords": keywords,
                    "advice": advice_raw,
                }
            )

        return rules_local

    except Exception as e:
        messagebox.showerror(
            "에러",
            f"규칙 파일을 읽는 중 오류가 발생했습니다:\n{e}"
        )
        return []


def save_rules_to_xlsx(rules_local, path=XLSX_FILE):
    """현재 rules 리스트를 통째로 rules.xlsx에 저장"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "rules"

        headers = ["category", "name", "keywords", "advice"]
        for col, h in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=h)

        for row_idx, r in enumerate(rules_local, start=2):
            ws.cell(row=row_idx, column=1, value=r["category"])
            ws.cell(row=row_idx, column=2, value=r["name"])
            ws.cell(row=row_idx, column=3, value=",".join(r["keywords"]))
            ws.cell(row=row_idx, column=4, value=r["advice"])

        wb.save(path)
        status_var.set(f"{path} 저장 완료 (규칙 {len(rules_local)}개)")
    except Exception as e:
        messagebox.showerror(
            "에러",
            f"규칙을 엑셀에 저장하는 중 오류가 발생했습니다:\n{e}"
        )


# ================= 필터링 로직 ================= #

def get_categories(rules_local):
    cats = sorted(
        set(r["category"] for r in rules_local),
        key=lambda x: x.lower()
    )
    return ["전체"] + cats


def filter_rules_by_category(rules_local, category):
    if not rules_local:
        return []
    if category in (None, "", "전체"):
        return rules_local[:]
    return [r for r in rules_local if r["category"] == category]


def search_rules(rules_local, query):
    q = query.strip().lower()
    if not q:
        return rules_local[:]
    matched = []
    for r in rules_local:
        text_blob = " ".join([
            r["name"],
            " ".join(r["keywords"]),
            r["advice"],
        ]).lower()
        if q in text_blob:
            matched.append(r)
    return matched


# ================= UI 동작 함수 ================= #

def refresh_category_list():
    """왼쪽 카테고리 리스트 갱신"""
    category_listbox.delete(0, tk.END)
    for cat in categories:
        category_listbox.insert(tk.END, cat)
    if categories:
        category_listbox.select_set(0)
    on_category_select(None)


def refresh_rule_list(display_list):
    global displayed_rules

    # 규칙을 name 기준으로 가나다/ABC 순 정렬
    displayed_rules = sorted(display_list, key=lambda r: r["name"].lower())

    rule_listbox.delete(0, tk.END)
    for r in displayed_rules:
        rule_listbox.insert(tk.END, f"[{r['category']}] {r['name']}")

    advice_text.config(state="normal")
    advice_text.delete("1.0", tk.END)
    advice_text.config(state="disabled")

    status_var.set(f"규칙 {len(displayed_rules)}개 (총 {len(rules)}개 중)")


def on_category_select(event):
    if not rules:
        return

    sel = category_listbox.curselection()
    if not sel:
        return

    cat = category_listbox.get(sel[0])
    base = filter_rules_by_category(rules, cat)

    query = search_var.get()
    filtered = search_rules(base, query) if query.strip() else base
    refresh_rule_list(filtered)


def on_rule_select(event):
    if not displayed_rules:
        return

    sel = rule_listbox.curselection()
    if not sel:
        return

    idx = sel[0]
    if idx < 0 or idx >= len(displayed_rules):
        return

    rule = displayed_rules[idx]

    advice_text.config(state="normal")
    advice_text.delete("1.0", tk.END)
    advice_text.insert(tk.END, f"[{rule['category']}] {rule['name']}\n\n")
    advice_text.insert(tk.END, strip_sources(rule["advice"]))
    advice_text.config(state="disabled")

# ================= Stage → 규칙 연동 ================= #

# Stage 결과 문자열(예: "Stage IA1")을 rules.xlsx의 name과 매핑합니다.
# 기본은 동일한 문자열을 name으로 찾고, 없으면 일부 대체 규칙명을 시도합니다.
STAGE_TO_RULE_NAME = {
    "Stage 0": "Stage 0",
    "Stage IA1": "Stage IA1",
    "Stage IA2": "Stage IA2",
    "Stage IA3": "Stage IA3",
    "Stage IB": "Stage IB",
    "Stage IIA": "Stage IIA",
    "Stage IIB": "Stage IIB",
    "Stage IIIA": "Stage IIIA",
    "Stage IIIB": "Stage IIIB",
    "Stage IIIC": "Stage IIIC",
    "Stage IVA": "Stage IVA",
    "Stage IVB": "Stage IVB",
    # range 표현(선택 사항): rules.xlsx에 해당 name이 있으면 열립니다.
    "Stage I–III (T에 따라)": "Stage I–III (T에 따라)",
    "Stage IIB–IIIA": "Stage IIB–IIIA",
    "Stage IIIA–IIIB": "Stage IIIA–IIIB",
    "Stage IIIB–IIIC": "Stage IIIB–IIIC",
}

def open_rule_by_stage(stage_text: str):
    """Stage 결과에 맞는 규칙을 자동으로 찾아 선택하고 advice를 표시"""
    stage_text = (stage_text or "").strip()
    if not stage_text:
        messagebox.showinfo("안내", "Stage 결과가 비어 있습니다.")
        return

    target_name = STAGE_TO_RULE_NAME.get(stage_text, stage_text)

    # 1) Stage는 특정 카테고리에 한정하지 않고 '전체'에서 검색
    #    (사용자가 rules.xlsx에서 카테고리를 어떻게 두었든 name만 맞으면 찾도록)
    try:
        idx_all = None
        for i in range(category_listbox.size()):
            if category_listbox.get(i).strip() == "전체":
                idx_all = i
                break
        if idx_all is None:
            idx_all = 0  # 안전장치: 첫 항목
        category_listbox.select_clear(0, tk.END)
        category_listbox.select_set(idx_all)
        category_listbox.see(idx_all)
        on_category_select(None)
    except Exception:
        pass

    # 2) displayed_rules에서 name 매칭(정확히 우선, 없으면 포함 검색)
    if not displayed_rules:
        return

    found_idx = None
    for i, r in enumerate(displayed_rules):
        if (r.get("name") or "").strip() == target_name:
            found_idx = i
            break
    if found_idx is None:
        for i, r in enumerate(displayed_rules):
            if target_name.lower() in (r.get("name") or "").lower():
                found_idx = i
                break

    if found_idx is None:
        messagebox.showinfo("안내", f"'{stage_text}'에 해당하는 규칙(name='{target_name}')을(를) 찾지 못했습니다.\n"
                                   f"rules.xlsx에 동일한 name으로 항목을 추가해 주세요.")
        return

    rule_listbox.select_clear(0, tk.END)
    rule_listbox.select_set(found_idx)
    rule_listbox.see(found_idx)
    on_rule_select(None)




def open_rule_by_name(rule_name: str, prefer_category: str | None = None):
    """name으로 규칙을 찾아 선택/표시.
    prefer_category가 주어지면 해당 카테고리를 먼저 선택 후 검색합니다.
    """
    rule_name = (rule_name or "").strip()
    if not rule_name:
        messagebox.showinfo("안내", "열 규칙 이름(name)이 비어 있습니다.")
        return

    # 1) 카테고리 선택: prefer_category가 있으면 우선, 없으면 '전체'
    target_cat = prefer_category or "전체"
    try:
        idx_cat = None
        for i in range(category_listbox.size()):
            if category_listbox.get(i).strip() == target_cat:
                idx_cat = i
                break
        if idx_cat is None:
            # fallback: 전체
            for i in range(category_listbox.size()):
                if category_listbox.get(i).strip() == "전체":
                    idx_cat = i
                    break
        if idx_cat is None:
            idx_cat = 0
        category_listbox.select_clear(0, tk.END)
        category_listbox.select_set(idx_cat)
        category_listbox.see(idx_cat)
        on_category_select(None)
    except Exception:
        pass

    if not displayed_rules:
        return

    # 2) name 정확 일치 우선, 없으면 포함 검색
    found_idx = None
    for i, r in enumerate(displayed_rules):
        if (r.get("name") or "").strip() == rule_name:
            found_idx = i
            break
    if found_idx is None:
        for i, r in enumerate(displayed_rules):
            if rule_name.lower() in (r.get("name") or "").lower():
                found_idx = i
                break

    if found_idx is None:
        messagebox.showinfo("안내", f"규칙(name='{rule_name}')을(를) 찾지 못했습니다.\n"
                                   f"rules.xlsx에 동일한 name으로 항목이 있는지 확인해 주세요.")
        return

    rule_listbox.select_clear(0, tk.END)
    rule_listbox.select_set(found_idx)
    rule_listbox.see(found_idx)
    on_rule_select(None)

def on_search():
    if not rules:
        return

    query = search_var.get()
    sel = category_listbox.curselection()
    cat = None
    if sel:
        cat = category_listbox.get(sel[0])

    base = filter_rules_by_category(rules, cat)
    filtered = search_rules(base, query)
    refresh_rule_list(filtered)


def on_reload_rules():
    """엑셀 다시 읽기"""
    global rules, categories
    rules = load_rules_from_xlsx()
    categories = get_categories(rules) if rules else ["전체"]
    refresh_category_list()
    status_var.set("엑셀에서 다시 불러왔습니다.")


def on_save_to_excel():
    save_rules_to_xlsx(rules)


# ================= 규칙 편집 창 ================= #

def open_rule_editor(edit_mode="edit"):
    """선택 규칙 수정 또는 새 규칙 추가를 위한 팝업"""
    if edit_mode == "edit":
        if not displayed_rules:
            messagebox.showwarning("알림", "수정할 규칙을 먼저 선택해주세요.")
            return
        sel = rule_listbox.curselection()
        if not sel:
            messagebox.showwarning("알림", "수정할 규칙을 먼저 선택해주세요.")
            return
        idx = sel[0]
        rule = displayed_rules[idx]
        existing_index = rules.index(rule)
    else:  # new
        rule = {"category": "", "name": "", "keywords": [], "advice": ""}
        existing_index = None

    editor = tk.Toplevel(root)
    editor.title("규칙 편집" if edit_mode == "edit" else "새 규칙 추가")
    editor.geometry("600x500")

    editor.columnconfigure(1, weight=1)
    for r_i in range(4):
        editor.rowconfigure(r_i, weight=0)
    editor.rowconfigure(3, weight=1)

    # category
    tk.Label(editor, text="category").grid(row=0, column=0, padx=5, pady=5, sticky="e")
    cat_var = tk.StringVar(value=rule["category"])
    tk.Entry(editor, textvariable=cat_var).grid(row=0, column=1, padx=5, pady=5, sticky="ew")

    # name
    tk.Label(editor, text="name").grid(row=1, column=0, padx=5, pady=5, sticky="e")
    name_var = tk.StringVar(value=rule["name"])
    tk.Entry(editor, textvariable=name_var).grid(row=1, column=1, padx=5, pady=5, sticky="ew")

    # keywords
    tk.Label(editor, text="keywords(쉼표로 구분)").grid(row=2, column=0, padx=5, pady=5, sticky="e")
    kw_var = tk.StringVar(value=",".join(rule["keywords"]))
    tk.Entry(editor, textvariable=kw_var).grid(row=2, column=1, padx=5, pady=5, sticky="ew")

    # advice
    tk.Label(editor, text="advice").grid(row=3, column=0, padx=5, pady=5, sticky="ne")
    adv_box = scrolledtext.ScrolledText(editor, wrap="word")
    adv_box.grid(row=3, column=1, padx=5, pady=5, sticky="nsew")
    adv_box.insert("1.0", rule["advice"])

    def save_and_close():
        new_cat = cat_var.get().strip()
        new_name = name_var.get().strip()
        new_kw_raw = kw_var.get().strip()
        new_adv = adv_box.get("1.0", tk.END).rstrip("\n")

        if not new_cat or not new_name:
            messagebox.showwarning("알림", "category와 name은 필수입니다.")
            return

        new_rule = {
            "category": new_cat,
            "name": new_name,
            "keywords": [
                kw.strip()
                for kw in new_kw_raw.split(",")
                if kw.strip()
            ],
            "advice": new_adv,
        }

        if existing_index is None:
            rules.append(new_rule)
        else:
            rules[existing_index] = new_rule

        # 카테고리/리스트 갱신
        global categories
        categories = get_categories(rules)
        refresh_category_list()
        editor.destroy()

    def delete_rule():
        if existing_index is None:
            editor.destroy()
            return
        if messagebox.askyesno("삭제 확인", "이 규칙을 삭제하시겠습니까?"):
            del rules[existing_index]
            global categories
            categories = get_categories(rules) if rules else ["전체"]
            refresh_category_list()
            editor.destroy()

    btn_frame = tk.Frame(editor)
    btn_frame.grid(row=4, column=0, columnspan=2, pady=8)

    tk.Button(btn_frame, text="저장", command=save_and_close, width=10).pack(side="left", padx=5)
    if edit_mode == "edit":
        tk.Button(btn_frame, text="삭제", command=delete_rule, width=10).pack(side="left", padx=5)
    tk.Button(btn_frame, text="닫기", command=editor.destroy, width=10).pack(side="right", padx=5)


# ================= 메인 윈도우 / 레이아웃 ================= #


# ================= 외래 기록(SOAP) 작성 도구 ================= #

SOAP_TEMPLATES = [
    "호흡기 일반",
    "ILD 초진",
    "ILD 재진",
    "폐암",
    "COPD",
]

def _match_template(template_name: str, rule_category: str) -> bool:
    """템플릿명과 rule category를 느슨하게 매칭합니다."""
    if not template_name:
        return True
    if not rule_category:
        return False
    t = str(template_name).strip().lower()
    c = str(rule_category).strip().lower()

    # 정확히 같으면 OK
    if t == c:
        return True

    # 핵심 키워드 매칭 (ILD 초진/재진 등)
    # 예: template "ILD 초진" -> 키워드 ["ild", "초진"]
    keywords = [k for k in re.split(r"\s+", t) if k]
    # 영어/약어 기반으로도 매칭
    if "ild" in t and "ild" in c:
        pass
    # 모두 포함되면 매칭
    if all(k in c for k in keywords):
        return True

    # 부분 포함(호흡기 일반은 '호흡기'만 포함해도)
    if t in c or c in t:
        return True

    # COPD/결핵/폐암은 포함 기반
    for key in ["copd", "결핵", "tb", "폐암", "lung cancer", "asthma", "천식", "ild", "interstitial"]:
        if key.lower() in t and key.lower() in c:
            return True

    return False


class OutpatientNoteWindow(tk.Toplevel):
    """템플릿별로 규칙을 선택해 SOAP 텍스트를 생성/복사하는 창"""

    def __init__(self, master, template_name: str, rules_all: list):
        super().__init__(master)
        self.title(f"외래 기록 작성 - {template_name}")
        self.geometry("980x620")
        self.minsize(900, 560)

        self.template_name = template_name
        self.rules_all = rules_all or []

        # 템플릿에 맞는 규칙 후보
        matched = [r for r in self.rules_all if _match_template(template_name, r.get("category", ""))]
        self.rules = matched if matched else self.rules_all[:]  # 매칭이 없으면 전체를 보여줌

        self.filtered = self.rules[:]

        self.soap = {"S": [], "O": [], "A": [], "P": []}

        # --- 레이아웃: 좌(규칙) / 우(SOAP) ---
        paned = ttk.Panedwindow(self, orient="horizontal")
        paned.pack(fill="both", expand=True, padx=8, pady=8)

        left = ttk.Frame(paned)
        right = ttk.Frame(paned)
        paned.add(left, weight=1)
        paned.add(right, weight=1)

        # ===== Left: 검색 + 규칙 리스트 =====
        ttk.Label(left, text="규칙 목록", font=("맑은 고딕", 10, "bold")).pack(anchor="w")

        search_row = ttk.Frame(left)
        search_row.pack(fill="x", pady=(6, 6))
        ttk.Label(search_row, text="검색").pack(side="left")
        self.search_var = tk.StringVar()
        search_entry = ttk.Entry(search_row, textvariable=self.search_var)
        search_entry.pack(side="left", fill="x", expand=True, padx=(6, 6))
        ttk.Button(search_row, text="적용", command=self.apply_filter).pack(side="left")
        ttk.Button(search_row, text="초기화", command=self.reset_filter).pack(side="left", padx=(6, 0))
        search_entry.bind("<Return>", lambda e: self.apply_filter())

        list_frame = ttk.Frame(left)
        list_frame.pack(fill="both", expand=True)
        self.rule_listbox = tk.Listbox(list_frame, height=18, exportselection=False)
        rule_scroll = ttk.Scrollbar(list_frame, command=self.rule_listbox.yview)
        self.rule_listbox.config(yscrollcommand=rule_scroll.set)
        self.rule_listbox.pack(side="left", fill="both", expand=True)
        rule_scroll.pack(side="right", fill="y")

        self.rule_listbox.bind("<<ListboxSelect>>", self.on_select_rule)
        self.rule_listbox.bind("<Double-Button-1>", lambda e: self.add_selected_to_section("S"))

        ttk.Label(left, text="미리보기", font=("맑은 고딕", 10, "bold")).pack(anchor="w", pady=(10, 2))
        self.preview = scrolledtext.ScrolledText(left, height=8, wrap="word", font=("맑은 고딕", 10))
        self.preview.pack(fill="both", expand=False)

        # ===== Right: SOAP 구성 =====
        header = ttk.Frame(right)
        header.pack(fill="x")
        ttk.Label(header, text="SOAP 구성", font=("맑은 고딕", 10, "bold")).pack(side="left")
        ttk.Button(header, text="전체 삭제", command=self.clear_all).pack(side="right")

        # 추가 버튼 줄
        add_row = ttk.Frame(right)
        add_row.pack(fill="x", pady=(8, 6))
        ttk.Label(add_row, text="선택 항목 추가 →").pack(side="left")
        for sec in ["S", "O", "A", "P"]:
            ttk.Button(add_row, text=sec, command=lambda s=sec: self.add_selected_to_section(s)).pack(side="left", padx=4)

        # 섹션별 리스트
        sec_frame = ttk.Frame(right)
        sec_frame.pack(fill="both", expand=True)

        self.sec_boxes = {}
        for i, sec in enumerate(["S", "O", "A", "P"]):
            col = i % 2
            row = i // 2
            box = ttk.LabelFrame(sec_frame, text=sec)
            box.grid(row=row, column=col, sticky="nsew", padx=6, pady=6)
            sec_frame.columnconfigure(col, weight=1)
            sec_frame.rowconfigure(row, weight=1)

            lb = tk.Listbox(box, exportselection=False)
            sb = ttk.Scrollbar(box, command=lb.yview)
            lb.config(yscrollcommand=sb.set)
            lb.pack(side="left", fill="both", expand=True)
            sb.pack(side="right", fill="y")

            # 삭제 버튼
            btns = ttk.Frame(box)
            btns.pack(fill="x", pady=(4, 0))
            ttk.Button(btns, text="선택 삭제", command=lambda s=sec: self.remove_selected(s)).pack(side="left")
            ttk.Button(btns, text="위로", command=lambda s=sec: self.move_item(s, -1)).pack(side="left", padx=4)
            ttk.Button(btns, text="아래로", command=lambda s=sec: self.move_item(s, +1)).pack(side="left")

            self.sec_boxes[sec] = lb

        # 출력 영역
        out_frame = ttk.LabelFrame(right, text="외래 기록 출력")
        out_frame.pack(fill="both", expand=False, padx=6, pady=(6, 0))

        self.output = scrolledtext.ScrolledText(out_frame, height=8, wrap="word", font=("맑은 고딕", 10))
        self.output.pack(fill="both", expand=True, padx=6, pady=6)

        out_btns = ttk.Frame(out_frame)
        out_btns.pack(fill="x", padx=6, pady=(0, 6))
        ttk.Button(out_btns, text="출력 갱신", command=self.refresh_output).pack(side="left")
        ttk.Button(out_btns, text="복사", command=self.copy_output).pack(side="left", padx=6)

        # 초기 목록 채우기
        self.populate_rule_list()
        self.refresh_output()

    def populate_rule_list(self):
        self.rule_listbox.delete(0, tk.END)
        # 이름순 정렬
        self.filtered.sort(key=lambda r: str(r.get("name", "")).lower())
        for r in self.filtered:
            name = r.get("name", "")
            cat = r.get("category", "")
            self.rule_listbox.insert(tk.END, f"[{cat}] {name}")

    def apply_filter(self):
        q = self.search_var.get().strip().lower()
        if not q:
            self.filtered = self.rules[:]
        else:
            def _hit(r):
                return (q in str(r.get("name", "")).lower()
                        or q in str(r.get("keywords", "")).lower()
                        or q in str(r.get("advice", "")).lower()
                        or q in str(r.get("category", "")).lower())
            self.filtered = [r for r in self.rules if _hit(r)]
        self.populate_rule_list()
        self.preview.delete("1.0", tk.END)

    def reset_filter(self):
        self.search_var.set("")
        self.filtered = self.rules[:]
        self.populate_rule_list()
        self.preview.delete("1.0", tk.END)

    def _get_selected_rule(self):
        sel = self.rule_listbox.curselection()
        if not sel:
            return None
        idx = sel[0]
        if idx < 0 or idx >= len(self.filtered):
            return None
        return self.filtered[idx]

    def on_select_rule(self, _evt=None):
        r = self._get_selected_rule()
        self.preview.delete("1.0", tk.END)
        if not r:
            return
        name = r.get("name", "")
        cat = r.get("category", "")
        kw = r.get("keywords", "")
        adv = r.get("advice", "")
        text = f"[{cat}] {name}\n\n키워드: {kw}\n\n{adv}"
        self.preview.insert(tk.END, text)

    def _format_snippet(self, r):
        """EMR에 붙여넣을 최소 문장: 현재는 name: advice 1줄 요약"""
        name = str(r.get("name", "")).strip()
        adv = str(r.get("advice", "")).strip()
        if not adv:
            return name
        # advice 첫 줄만 요약으로 사용 (너무 길어지는 것을 방지)
        first = adv.splitlines()[0].strip()
        if len(first) > 140:
            first = first[:140].rstrip() + "…"
        if name:
            return f"{name}: {first}"
        return first

    def add_selected_to_section(self, section: str):
        r = self._get_selected_rule()
        if not r:
            return
        snippet = self._format_snippet(r)
        self.soap[section].append(snippet)
        self.sec_boxes[section].insert(tk.END, snippet)
        self.refresh_output()

    def remove_selected(self, section: str):
        lb = self.sec_boxes[section]
        sel = lb.curselection()
        if not sel:
            return
        idx = sel[0]
        lb.delete(idx)
        if 0 <= idx < len(self.soap[section]):
            self.soap[section].pop(idx)
        self.refresh_output()

    def move_item(self, section: str, delta: int):
        lb = self.sec_boxes[section]
        sel = lb.curselection()
        if not sel:
            return
        i = sel[0]
        j = i + delta
        if j < 0 or j >= lb.size():
            return
        # swap in data
        self.soap[section][i], self.soap[section][j] = self.soap[section][j], self.soap[section][i]
        # refresh listbox
        items = self.soap[section]
        lb.delete(0, tk.END)
        for it in items:
            lb.insert(tk.END, it)
        lb.selection_set(j)
        self.refresh_output()

    def clear_all(self):
        for sec in ["S", "O", "A", "P"]:
            self.soap[sec] = []
            self.sec_boxes[sec].delete(0, tk.END)
        self.refresh_output()

    def refresh_output(self):
        parts = []
        for sec in ["S", "O", "A", "P"]:
            if self.soap[sec]:
                lines = "\n".join(f"- {x}" for x in self.soap[sec])
            else:
                lines = ""
            parts.append(f"{sec}:\n{lines}".rstrip())
        text = "\n\n".join(parts).strip() + "\n"
        self.output.delete("1.0", tk.END)
        self.output.insert(tk.END, text)

    def copy_output(self):
        self.refresh_output()
        text = self.output.get("1.0", "end-1c")
        self.clipboard_clear()
        self.clipboard_append(text)
        messagebox.showinfo("복사됨", "외래 기록 텍스트를 클립보드에 복사했습니다.")




# ================= 외래 기록(체크리스트) 작성 도구 ================= #

# 1) 공통(호흡기 일반) 현재 증상 + mMRC
SYMPTOMS_BASE_YN = [
    "Cough (기침)",
    "Sputum (가래)",
    "Rhinorrhea (콧물)",
    "Nasal congestion (코막힘)",
    "Sore throat (인후통)",
    "Hemoptysis (혈담/객혈)",
    "Dyspnea (호흡곤란, at rest / exertional)",
    "Wheezing (천명음)",
    "Chest pain or tightness (흉통/답답함)",
    "Fever/chill (발열/오한)",
    "Fatigue (피로)",
    "Weight loss (체중 감소)",
    "현재 흡연 중",
]

# 2) ILD 재진 추가 문항
ILD_FU_YN = [
    "Dry, persistent cough (건성·지속 기침)",
    "Progressive dyspnea (점진적 악화 호흡곤란)",
    "Orthopnea/PND (기좌호흡/야간발작호흡곤란)",
]

# 3) ILD 초진 추가 문항
ILD_NEW_YN = [
    "증상 시작 후 기간 6개월 이상",
    "점진적 호흡곤란/건성기침 지속",
    "체중 감소/식욕 저하(최근 6개월)",
    "손가락지팡이증/관절통/피부 변화",
    "가족력(ILD/IPF/CTD) 있음",
    "흡연력(현재/과거 팩-이어) 있음",
    "직업/환경 노출(석면/분진/조류/농약) 있음",
    "약제 노출(암약제/면역억제제/항암제) 있음",
    "CTD 증상(레이노/관절염/근력 저하/피부 경화) 있음",
    "동반질환(고혈압/GERD/당뇨/심부전) 있음",
    "이전 PFT(6MWT) 결과 악화 추세",
    "급성 악화(AE) 병력 있음",
]

# 4) 폐암 환자용 추가 문항
LUNG_CA_YN = [
    "최근 수주–수개월간 기침/호흡곤란 악화",
    "객혈 또는 혈담 경험",
    "설명되지 않는 체중 감소/식욕 저하",
    "흉통/어깨통증/뼈 통증",
    "쉰목소리(hoarseness) 또는 삼킴 곤란",
    "두통, 어지러움, 신경학적 증상(마비, 감각저하 등)",
    "이전 항암/방사선 치료력 있음",
    "마지막 치료 후 악화된 증상 있음",
    "현재 복용 중 항암제/표적치료제/면역항암제 있음",
    "관련 의심 부작용(피부, 호흡, 위장, 간/신장 기능 이상 등) 있음",
]

# 5) COPD 환자용 추가 문항
COPD_YN = [
    "만성 기침(3개월 이상/2년 이상 반복)",
    "만성 가래(아침/계절성 포함)",
    "최근 1년간 악화(Exacerbation)로 응급실/입원 경험",
    "계단 오르기/평지 보행 시 숨참 증가",
    "야간/새벽에 악화되는 호흡곤란",
    "흡연 중이거나 과거 흡연력(팩-이어) 있음",
    "직업적/환경적 노출(분진, 화학물질, 실내·실외 공기오염) 있음",
    "흡입제(ICS/LABA/LAMA 등) 규칙적 사용",
    "흡입기 사용법 교육 받은 적 있음",
    "최근 흡입제 종류 변경/중단 있음",
    "산소치료(가정 산소) 사용 중",
    "처방된 시간/유량대로 사용함",
    "흡입제/산소 사용과 관련된 불편감/부작용 있음",
    "기타 호흡기 질환 공통 질문 생활습관 및 위험인자",
]

# 6) 약제 사용 및 부작용
MED_AE_YN = [
    "약제 처방대로 복용/사용함",
    "경구 항생제 복용함",
    "최근 약제 중단/용량 변경 있음",
    "스테로이드 부작용(부종, 고혈당, 불면 등) 있음",
    "항섬유화제 부작용(GI, 피부, 간기능 등) 있음",
    "흡입제 사용법 숙지/규칙적 사용",
    "흡입제 부작용(구강칸디다, 쉰목소리 등) 있음",
]

# 7) Lab / 영상 검사 결과
LAB_IMG_YN = [
    "CBC abnormality",
    "BUN/Cr abnl",
    "OT/PT/T.bil elevation",
    "e' abnl",
    "CXR abnl",
    "CT result",
    "PFT FEV1/FVC/Ratio ///",
    "DLco",
    "5MWT",
]

# 8) Further Plan
PLAN_YN = [
    "Add",
    "d/c",
    "다음 PFT 검사",
    "다음 CT 검사",
    "입원",
]

# 특수 항목 키(라디오 그룹)
_SPECIAL_MMRC = "__MMRC__"
_SPECIAL_OPDFU = "__OPDFU__"

OPD_FU_CHOICES = [
    ("1wk", "1wk"),
    ("2wk", "2wk"),
    ("1m", "1m"),
    ("3m", "3m"),
    ("6m", "6m"),
]

# 템플릿 → 탭 구성
# 각 탭: (탭 제목, yn_items(list[str]), include_mmrc(bool), include_opdfu(bool))
TEMPLATE_TO_TABS = {
    "호흡기 일반": [
        ("현재 증상", SYMPTOMS_BASE_YN, True, False),
        ("약제/부작용", MED_AE_YN, False, False),
        ("Lab/영상", LAB_IMG_YN, False, False),
        ("Further Plan", PLAN_YN, False, True),
    ],
    "ILD 초진": [
        ("현재 증상", SYMPTOMS_BASE_YN, True, False),
        ("ILD 초진 추가", ILD_NEW_YN, False, False),
        ("약제/부작용", MED_AE_YN, False, False),
        ("Lab/영상", LAB_IMG_YN, False, False),
        ("Further Plan", PLAN_YN, False, True),
    ],
    "ILD 재진": [
        ("현재 증상", SYMPTOMS_BASE_YN, True, False),
        ("ILD 재진 추가", ILD_FU_YN, False, False),
        ("약제/부작용", MED_AE_YN, False, False),
        ("Lab/영상", LAB_IMG_YN, False, False),
        ("Further Plan", PLAN_YN, False, True),
    ],
    "폐암": [
        ("현재 증상", SYMPTOMS_BASE_YN, True, False),
        ("폐암 추가", LUNG_CA_YN, False, False),
        ("약제/부작용", MED_AE_YN, False, False),
        ("Lab/영상", LAB_IMG_YN, False, False),
        ("Further Plan", PLAN_YN, False, True),
    ],
    "COPD": [
        ("현재 증상", SYMPTOMS_BASE_YN, True, False),
        ("COPD 추가", COPD_YN, False, False),
        ("약제/부작용", MED_AE_YN, False, False),
        ("Lab/영상", LAB_IMG_YN, False, False),
        ("Further Plan", PLAN_YN, False, True),
    ],
}


class OutpatientChecklistWindow(tk.Toplevel):
    """템플릿별 Yes/No(또는 특수 라디오) 체크 후 외래 기록용 텍스트를 생성하는 창"""

    def __init__(self, master, template_name: str):
        super().__init__(master)
        self.template_name = template_name
        self.title(f"외래 기록 작성 - {template_name}")
        self.geometry("900x820")
        self.minsize(820, 760)

        self.tabs = TEMPLATE_TO_TABS.get(template_name, TEMPLATE_TO_TABS["호흡기 일반"])

        # yn item -> IntVar (-1 blank, 1 yes, 0 no)
        self.vars: dict[str, tk.IntVar] = {}
        # special vars
        self.special_vars = {
            _SPECIAL_MMRC: tk.IntVar(value=-1),
            _SPECIAL_OPDFU: tk.StringVar(value=""),
        }

        top = ttk.Frame(self)
        top.pack(fill="x", padx=12, pady=10)
        ttk.Label(top, text=f"{template_name} 문진 체크", font=("맑은 고딕", 12, "bold")).pack(side="left")

        # Notebook (tabs)
        self.nb = ttk.Notebook(self)
        self.nb.pack(fill="both", expand=True, padx=12, pady=(0, 10))

        self._build_tabs()
        self._bind_tab_shortcuts()

        # Output (항상 유지)
        out_box = ttk.LabelFrame(self, text="외래 기록 출력 (복사해서 EMR에 붙여넣기)")
        out_box.pack(fill="both", expand=False, padx=12, pady=(0, 12))

        self.output = scrolledtext.ScrolledText(out_box, height=10, wrap="word")
        self.output.pack(fill="both", expand=True, padx=10, pady=10)

        # Buttons
        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=12, pady=(0, 12))
        ttk.Button(btns, text="복사", command=self.copy_to_clipboard).pack(side="left")
        ttk.Button(btns, text="초기화", command=self.reset_all).pack(side="left", padx=(8, 0))
        ttk.Button(btns, text="닫기", command=self.destroy).pack(side="right")

        self.generate_output()

    def _build_tabs(self):
        for tab_title, yn_items, include_mmrc, include_opdfu in self.tabs:
            tab = ttk.Frame(self.nb)
            self.nb.add(tab, text=tab_title)

            # Scrollable frame inside each tab
            holder = ttk.Frame(tab)
            holder.pack(fill="both", expand=True)

            canvas = tk.Canvas(holder, highlightthickness=0)
            vbar = ttk.Scrollbar(holder, orient="vertical", command=canvas.yview)
            inner = ttk.Frame(canvas)

            inner.bind("<Configure>", lambda e, c=canvas: c.configure(scrollregion=c.bbox("all")))
            canvas.create_window((0, 0), window=inner, anchor="nw")
            canvas.configure(yscrollcommand=vbar.set)

            canvas.pack(side="left", fill="both", expand=True)
            vbar.pack(side="right", fill="y")

            def _on_mousewheel(event, c=canvas):
                c.yview_scroll(int(-1 * (event.delta / 120)), "units")
            canvas.bind("<Enter>", lambda _e, c=canvas: c.bind_all("<MouseWheel>", _on_mousewheel))
            canvas.bind("<Leave>", lambda _e, c=canvas: c.unbind_all("<MouseWheel>"))

            # Build rows
            r = 0
            if tab_title == "현재 증상":
                ttk.Label(inner, text="현재 증상 (Symptoms)", font=("맑은 고딕", 10, "bold")).grid(row=r, column=0, sticky="w", pady=(6, 6))
                r += 1

            for item in yn_items:
                self._add_yn_row(inner, r, item)
                r += 1

            if include_mmrc:
                ttk.Separator(inner).grid(row=r, column=0, sticky="ew", pady=(10, 10))
                r += 1
                self._add_mmrc_row(inner, r)
                r += 1

            if include_opdfu:
                ttk.Separator(inner).grid(row=r, column=0, sticky="ew", pady=(10, 10))
                r += 1
                self._add_opdfu_row(inner, r)
                r += 1

            inner.columnconfigure(0, weight=1)

    def _bind_tab_shortcuts(self):
        """Ctrl+1..5 로 탭 전환 (템플릿별 탭 개수만큼만 동작)"""
        def _go(idx: int):
            try:
                if 0 <= idx < len(self.nb.tabs()):
                    self.nb.select(idx)
            except Exception:
                pass

        for i in range(1, 6):
            # 두 표기 모두 바인딩(환경에 따라 다름)
            self.bind_all(f"<Control-Key-{i}>", lambda e, k=i: _go(k-1))
            self.bind_all(f"<Control-{i}>", lambda e, k=i: _go(k-1))

    def _add_yn_row(self, parent, row_idx: int, label: str):
        row = ttk.Frame(parent)
        row.grid(row=row_idx, column=0, sticky="ew", pady=2)
        row.columnconfigure(0, weight=1)

        ttk.Label(row, text=label).grid(row=0, column=0, sticky="w")

        v = self.vars.get(label)
        if v is None:
            v = tk.IntVar(value=-1)
            self.vars[label] = v
            v.trace_add('write', lambda *_: self.generate_output())

        # Yes/No/Blank
        btns = ttk.Frame(row)
        btns.grid(row=0, column=1, sticky="e")
        ttk.Radiobutton(btns, text="Yes", variable=v, value=1).pack(side="left", padx=6)
        ttk.Radiobutton(btns, text="No", variable=v, value=0).pack(side="left", padx=6)
        ttk.Radiobutton(btns, text="(빈칸)", variable=v, value=-1).pack(side="left", padx=6)

    def _add_mmrc_row(self, parent, row_idx: int):
        row = ttk.Frame(parent)
        row.grid(row=row_idx, column=0, sticky="ew", pady=2)
        row.columnconfigure(0, weight=1)

        ttk.Label(row, text="mMRC").grid(row=0, column=0, sticky="w")

        v = self.special_vars[_SPECIAL_MMRC]
        btns = ttk.Frame(row)
        btns.grid(row=0, column=1, sticky="e")
        for n in [0, 1, 2, 3, 4]:
            ttk.Radiobutton(btns, text=str(n), variable=v, value=n).pack(side="left", padx=4)
        ttk.Radiobutton(btns, text="(빈칸)", variable=v, value=-1).pack(side="left", padx=6)

        v.trace_add('write', lambda *_: self.generate_output())

    def _add_opdfu_row(self, parent, row_idx: int):
        row = ttk.Frame(parent)
        row.grid(row=row_idx, column=0, sticky="ew", pady=2)
        row.columnconfigure(0, weight=1)

        ttk.Label(row, text="OPD f/u").grid(row=0, column=0, sticky="w")

        v = self.special_vars[_SPECIAL_OPDFU]
        btns = ttk.Frame(row)
        btns.grid(row=0, column=1, sticky="e")
        for _disp, val in OPD_FU_CHOICES:
            ttk.Radiobutton(btns, text=_disp, variable=v, value=val).pack(side="left", padx=4)
        ttk.Radiobutton(btns, text="(빈칸)", variable=v, value="").pack(side="left", padx=6)

        v.trace_add('write', lambda *_: self.generate_output())

    @staticmethod
    def _strip_korean_parentheses(text: str) -> str:
        """문자열에서 '(한글 ...)' 괄호 부분만 제거합니다. (영문 설명 괄호는 유지)"""
        return re.sub(r"\s*\([^)]*[가-힣][^)]*\)", "", text).strip()

    def generate_output(self):
        # 출력 규칙:
        # - 빈칸(-1)은 출력에서 제외
        # - Yes(1)면 '+', No(0)면 '-'를 붙여 출력
        # - 각 탭(섹션)은 선택된 항목을 한 줄로 ', ' 연결
        lines = []

        for tab_title, yn_items, include_mmrc, include_opdfu in self.tabs:
            chosen = []

            for item in yn_items:
                val = self.vars[item].get()
                if val == -1:
                    continue

                cleaned = self._strip_korean_parentheses(item)
                if not cleaned:
                    continue

                if val == 1:
                    chosen.append(f"{cleaned} +")
                elif val == 0:
                    chosen.append(f"{cleaned} -")

            if include_mmrc:
                mmrc = self.special_vars[_SPECIAL_MMRC].get()
                if mmrc != -1:
                    chosen.append(f"mMRC {mmrc}")

            if include_opdfu:
                fu = self.special_vars[_SPECIAL_OPDFU].get()
                if str(fu).strip():
                    chosen.append(f"OPD f/u {fu}")

            if chosen:
                lines.append(", ".join(chosen))
                lines.append("")

        base_text = "\n".join(lines).rstrip()

        # 고정 문구는 항상 마지막에 추가
        if base_text:
            text_out = base_text + "\n\n" + FINAL_OUTPATIENT_NOTE + "\n"
        else:
            text_out = FINAL_OUTPATIENT_NOTE + "\n"

        self.output.delete("1.0", "end")
        self.output.insert("1.0", text_out)

    def copy_to_clipboard(self):
        text_out = self.output.get("1.0", "end-1c")
        self.clipboard_clear()
        self.clipboard_append(text_out)
        messagebox.showinfo("복사됨", "외래 기록 텍스트를 클립보드에 복사했습니다.")

    def reset_all(self):
        for v in self.vars.values():
            v.set(-1)
        self.special_vars[_SPECIAL_MMRC].set(-1)
        self.special_vars[_SPECIAL_OPDFU].set("")
        self.generate_output()
class OutpatientTemplateChooser(tk.Toplevel):
    """템플릿 선택 창"""

    def __init__(self, master, rules_all: list):
        super().__init__(master)
        self.title("외래 기록 작성 - 템플릿 선택")
        self.geometry("420x360")
        self.minsize(380, 320)
        self.rules_all = rules_all or []

        ttk.Label(self, text="외래 기록 템플릿을 선택하세요", font=("맑은 고딕", 11, "bold")).pack(pady=(16, 10))

        body = ttk.Frame(self)
        body.pack(fill="both", expand=True, padx=14, pady=10)

        for tname in SOAP_TEMPLATES:
            ttk.Button(body, text=tname, command=lambda n=tname: self.open_template(n)).pack(fill="x", pady=6)

        ttk.Separator(self).pack(fill="x", padx=14, pady=10)
        ttk.Button(self, text="닫기", command=self.destroy).pack(pady=(0, 12))

    def open_template(self, template_name: str):
        # 템플릿을 고르면 선택 창은 자동으로 닫히고, 체크리스트 창만 남게 합니다.
        OutpatientChecklistWindow(self.master, template_name=template_name)
        try:
            self.destroy()
        except Exception:
            pass


# ================= ILD 진단 알고리즘(조직검사/급성악화 포함) ================= #

ILD_RULE_MAP = {
    "IPF": "IPF 진단 치료",
    "HP": "HP 진단 치료",
    "CTD-ILD": "CTD-ILD 진단 치료",
    "iNSIP": "iNSIP 진단 치료",
    "DIP": "DIP 진단 치료",
    "COP": "COP 진단치료",
    "AFOP": "AFOP 진단 치료",
    "AIP": "AIP 진단 치료",
    "idiopathic LIP": "idiopathic LIP 진단 치료",
    "PPFE": "PPFE",
    "Unclassifiable IIP": "Unclassifiable IIP",
}


# AE(급성악화) 상황에서 우선 열 규칙(name) 후보들.
# rules.xlsx에 아래 name이 존재하면 순서대로 열기를 시도합니다.
AE_RULE_SEQUENCE = [
    "ILD 급성악화 – 감별진단",
    "ILD 급성악화 – 초기 처치",
    "ILD 급성악화 – 스테로이드",
    "ILD 급성악화 – 기계환기 판단",
    "ILD 급성악화 – 예후 및 목표 치료",
    "ILD 급성악화",  # 사용자가 단일 이름으로 만들어둔 경우 대비
]


class ILDAlgorithmWindow(tk.Toplevel):
    """체크 기반 ILD 진단 알고리즘 창 (조직검사 타이밍/급성악화 포함).
    - 결과가 나오면 해당 규칙(name) 창을 자동으로 열도록 설계
    """
    def __init__(self, master):
        super().__init__(master)
        self.title("ILD 진단 알고리즘")
        self.geometry("900x650")

        self.notebook = ttk.Notebook(self)
        # 레이아웃: 상단(탭) + 하단(출력)
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)
        self.rowconfigure(1, weight=0)

        self.notebook = ttk.Notebook(self)
        self.notebook.grid(row=0, column=0, sticky="nsew")

        self.tab_ae = ttk.Frame(self.notebook)
        self.tab_chronic = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_ae, text="1) 급성악화(AE)")
        self.notebook.add(self.tab_chronic, text="2) 만성 ILD 진단")

        # Ctrl+1/2로 탭 이동
        self._bind_ctrl_tab_shortcuts(self.notebook, 2)

        # ----- 하단 출력(모든 탭 공통) ----- #
        out_wrap = ttk.Frame(self)
        out_wrap.grid(row=1, column=0, sticky="nsew")
        out_wrap.columnconfigure(0, weight=1)
        out_wrap.rowconfigure(1, weight=1)

        ttk.Label(out_wrap, text="출력", font=("맑은 고딕", 10, "bold")).grid(row=0, column=0, sticky="w", padx=10, pady=(8, 4))

        self.output = scrolledtext.ScrolledText(out_wrap, height=12, wrap="word", font=("맑은 고딕", 10))
        self.output.grid(row=1, column=0, sticky="nsew", padx=10, pady=(0, 10))
        self.output.config(state="disabled")

        self._build_tab_ae()
        self._build_tab_chronic()

    def _set_text(self, widget: tk.Text, s: str):
        widget.config(state="normal")
        widget.delete("1.0", tk.END)
        widget.insert(tk.END, s)
        widget.config(state="disabled")

    # -------- Tab1: AE --------
    
    def _bind_ctrl_tab_shortcuts(self, notebook, n_tabs: int):
        for i in range(min(n_tabs, 9)):
            key = str(i + 1)

            def _handler(event, idx=i):
                try:
                    w = event.widget
                    if w is None or w.winfo_toplevel() != self:
                        return
                    notebook.select(idx)
                    return "break"
                except Exception:
                    return

            self.bind_all(f"<Control-Key-{key}>", _handler, add="+")

    def _build_tab_ae(self):
        ttk.Label(self.tab_ae, text="급성악화(AE) 평가", font=("맑은 고딕", 11, "bold")).pack(anchor="w", pady=(8, 6))

        f = ttk.LabelFrame(self.tab_ae, text="AE-IPF(2016 개정) 핵심 요소(간이)")
        f.pack(fill="x", padx=10, pady=6)

        self.ae_sudden30 = tk.BooleanVar(value=False)
        self.ae_known_ipf = tk.BooleanVar(value=False)
        self.ae_new_bilat_ggo = tk.BooleanVar(value=False)
        self.ae_not_hf_overload = tk.BooleanVar(value=False)

        ttk.Checkbutton(f, text="최근 30일 이내 급성 악화(호흡곤란/저산소증 급격 악화)", variable=self.ae_sudden30).pack(anchor="w", padx=10, pady=2)
        ttk.Checkbutton(f, text="IPF가 이미 진단되었거나(또는 현재 IPF로 강력 의심)", variable=self.ae_known_ipf).pack(anchor="w", padx=10, pady=2)
        ttk.Checkbutton(f, text="CT: 기존 UIP(또는 섬유화) 위에 신규 양측성 GGO 및/또는 consolidation", variable=self.ae_new_bilat_ggo).pack(anchor="w", padx=10, pady=2)
        ttk.Checkbutton(f, text="심부전/수액과다로 완전히 설명되지 않음", variable=self.ae_not_hf_overload).pack(anchor="w", padx=10, pady=2)

        btnf = ttk.Frame(self.tab_ae)
        btnf.pack(anchor="w", padx=10, pady=(6, 4))
        ttk.Button(btnf, text="AE 판정", command=self._run_ae).pack(side="left")
        ttk.Button(btnf, text="→ 만성 진단(탭2)", command=lambda: self.notebook.select(self.tab_chronic)).pack(side="left", padx=(10,0))

        self._set_text(self.output, "체크 후 'AE 판정'을 누르시면 결과가 표시됩니다.")

    def _run_ae(self):
        checks = [self.ae_sudden30.get(), self.ae_known_ipf.get(), self.ae_new_bilat_ggo.get(), self.ae_not_hf_overload.get()]
        if sum(checks) >= 3 and self.ae_sudden30.get() and self.ae_new_bilat_ggo.get():
            # AE-IPF 가능
            self._set_text(self.output,
                           "✅ AE-IPF 의심(간이 기준)\n"
                           "- 최근 30일 이내 급성 악화 + CT 신규 양측 GGO/경화 + HF/수액과다로 설명 어려움\n\n"
                           "➡ IPF 진단/치료 규칙 창을 엽니다.")
            self.after(200, self._open_ae_rules)
        else:
            self._set_text(self.output,
                           "⚠ 급성악화(AE)는 의심될 수 있으나, AE-IPF 간이 기준을 충분히 만족하지 않습니다.\n"
                           "- 감염/혈전/심부전/약물/기타 ILD-AE를 함께 평가하세요.\n\n"
                           "만성 ILD 진단은 Tab2에서 진행하세요.")

    def _open_ae_rules(self):
        """AE(급성악화)로 판단되었을 때 규칙(name) 창을 자동으로 열어줍니다.

        - 사용자가 rules.xlsx에 만들어 둔 AE 전용 규칙을 우선 오픈
        - 이어서 IPF 진단/치료 규칙도 같이 열어 (AE-IPF 상황에서) 약제/추적 등 확인이 가능하게 합니다.
        - ILD 알고리즘 창 자체는 닫지 않습니다.
        """

        def _rule_exists(nm: str) -> bool:
            try:
                return any((r.get("name") or "").strip() == nm for r in (rules or []))
            except Exception:
                return False

        # 1) AE 전용 규칙: 존재하는 첫 항목을 오픈
        opened_any = False
        for nm in AE_RULE_SEQUENCE:
            if _rule_exists(nm):
                try:
                    open_rule_by_name(nm, prefer_category="ILD")
                except Exception:
                    open_rule_by_name(nm)
                opened_any = True
                break

        # 2) AE 전용 규칙이 하나도 없더라도, IPF 진단/치료는 열어 둡니다.
        ipf_rule = ILD_RULE_MAP.get("IPF", "IPF 진단 치료")
        if _rule_exists(ipf_rule):
            try:
                open_rule_by_name(ipf_rule, prefer_category="ILD")
            except Exception:
                open_rule_by_name(ipf_rule)

        # 3) 메인 창만 앞으로 올리고, 이 창은 유지
        try:
            self.master.lift()
            self.master.focus_force()
            self.master.attributes("-topmost", True)
            self.master.after(50, lambda: self.master.attributes("-topmost", False))
        except Exception:
            pass

    # -------- Tab2: Chronic --------
    def _build_tab_chronic(self):
        ttk.Label(self.tab_chronic, text="만성 ILD 진단(간단 트리 + 조직검사 타이밍 포함)", font=("맑은 고딕", 11, "bold")).pack(anchor="w", pady=(8, 6))

        # A. fibrotic?
        f1 = ttk.LabelFrame(self.tab_chronic, text="A. HRCT: Fibrotic vs Non-fibrotic")
        f1.pack(fill="x", padx=10, pady=6)
        self.fibrotic = tk.StringVar(value="fibrotic")
        ttk.Radiobutton(f1, text="Fibrotic ILD (reticulation/traction/honeycombing 등)", variable=self.fibrotic, value="fibrotic").pack(anchor="w", padx=10, pady=2)
        ttk.Radiobutton(f1, text="Non-fibrotic ILD (GGO/결절/침윤 우세)", variable=self.fibrotic, value="nonfibrotic").pack(anchor="w", padx=10, pady=2)

        # B. UIP pattern if fibrotic
        f2 = ttk.LabelFrame(self.tab_chronic, text="B. (Fibrotic ILD) HRCT UIP 패턴")
        f2.pack(fill="x", padx=10, pady=6)
        self.uip = tk.StringVar(value="indeterminate")
        ttk.Radiobutton(f2, text="UIP", variable=self.uip, value="uip").pack(anchor="w", padx=10, pady=2)
        ttk.Radiobutton(f2, text="Probable UIP", variable=self.uip, value="probable").pack(anchor="w", padx=10, pady=2)
        ttk.Radiobutton(f2, text="Indeterminate for UIP", variable=self.uip, value="indeterminate").pack(anchor="w", padx=10, pady=2)
        ttk.Radiobutton(f2, text="Alternative diagnosis", variable=self.uip, value="alternative").pack(anchor="w", padx=10, pady=2)

        # C. Etiology clues
        f3 = ttk.LabelFrame(self.tab_chronic, text="C. 원인 단서(있으면 해당 질환 우선)")
        f3.pack(fill="x", padx=10, pady=6)
        self.ctd_clue = tk.BooleanVar(value=False)
        self.hp_clue = tk.BooleanVar(value=False)
        self.smoking_related = tk.BooleanVar(value=False)
        self.op_pattern = tk.BooleanVar(value=False)
        self.dad_ards_like = tk.BooleanVar(value=False)
        self.lip_clue = tk.BooleanVar(value=False)
        self.ppfe_clue = tk.BooleanVar(value=False)

        ttk.Checkbutton(f3, text="CTD 단서(관절/피부/근염/레이노 등) 또는 자가항체 양성", variable=self.ctd_clue).pack(anchor="w", padx=10, pady=2)
        ttk.Checkbutton(f3, text="HP 단서(노출력+air-trapping/mosaic 등)", variable=self.hp_clue).pack(anchor="w", padx=10, pady=2)
        ttk.Checkbutton(f3, text="흡연 관련 ILD 의심(DIP 등)", variable=self.smoking_related).pack(anchor="w", padx=10, pady=2)
        ttk.Checkbutton(f3, text="OP 패턴(이동성 침윤/patchy consolidation 등) → COP 우선", variable=self.op_pattern).pack(anchor="w", padx=10, pady=2)
        ttk.Checkbutton(f3, text="급성 ARDS-like + DAD 의심 → AIP/AFOP 고려", variable=self.dad_ards_like).pack(anchor="w", padx=10, pady=2)
        ttk.Checkbutton(f3, text="LIP 단서(낭종+GGO, Sjogren/HIV 배제 후)", variable=self.lip_clue).pack(anchor="w", padx=10, pady=2)
        ttk.Checkbutton(f3, text="상엽 흉막하 우세 섬유화/반복 기흉 → PPFE 의심", variable=self.ppfe_clue).pack(anchor="w", padx=10, pady=2)

        btnf = ttk.Frame(self.tab_chronic)
        btnf.pack(anchor="w", padx=10, pady=(6, 4))
        ttk.Button(btnf, text="진단 도출", command=self._run_chronic).pack(side="left")

        self._set_text(self.output, "체크 후 '진단 도출'을 누르시면 결과가 표시됩니다.")

    def _diagnose_chronic(self):
        # 1) Strong etiology clues first
        if self.ctd_clue.get():
            return "CTD-ILD", "CTD 단서가 있어 CTD-ILD 우선 고려(자가항체/류마협진/MDD)."
        if self.hp_clue.get():
            return "HP", "노출력/air-trapping 등 HP 단서가 있어 HP 우선 고려(항원 회피/필요 시 BAL/조직)."
        if self.ppfe_clue.get():
            return "PPFE", "상엽 흉막하 우세 섬유화/pleural thickening → PPFE 의심."
        if self.lip_clue.get():
            return "idiopathic LIP", "LIP 단서(낭종+GGO). 이차 원인 배제 후 idiopathic LIP 고려."
        if self.op_pattern.get():
            return "COP", "OP 패턴(이동성/patchy consolidation). 이차 원인 배제 후 COP 고려."
        if self.dad_ards_like.get():
            # 급성 폐손상 쪽은 AIP/AFOP 분기: 여기서는 단순히 AIP로 안내
            return "AIP", "급성 ARDS-like + DAD 의심. 원인 배제 후 AIP/AFOP 감별(조직검사 고려)."
        if self.smoking_related.get():
            return "DIP", "흡연 관련 ILD(DIP 등) 가능. 금연 및 스테로이드 반응 가능."

        # 2) Fibrotic vs non-fibrotic + UIP pattern
        if self.fibrotic.get() == "fibrotic":
            if self.uip.get() in ("uip", "probable"):
                return "IPF", "Fibrotic ILD에서 UIP/Probable UIP이며 다른 원인 단서가 뚜렷하지 않아 IPF 가능성이 높습니다.\n- 다른 원인 배제 후: 조직검사 없이도 진단 가능할 수 있습니다."
            # Indeterminate/Alternative: 조직검사 핵심 타이밍
            return "iNSIP", ("UIP로 확정되지 않는 fibrotic ILD(Indeterminate/Alternative).\n"
                            "➡ NSIP/기타 IIP 감별이 필요하며 치료결정(면역억제 vs 항섬유화)에 중요하면 조직검사(TBLC/SLB)를 고려합니다.")
        else:
            # Non-fibrotic: NSIP 패턴 가능
            return "iNSIP", "Non-fibrotic ILD에서 NSIP 패턴 가능.\nCTD/HP/약물/감염 배제 후 필요 시 조직검사로 확진합니다."

    def _run_chronic(self):
        dx, reason = self._diagnose_chronic()
        msg = f"✅ 제안 진단: {dx}\n\n{reason}\n"
        # 조직검사 안내 강조
        if dx in ("iNSIP", "AIP") or (self.fibrotic.get()=="fibrotic" and self.uip.get() in ("indeterminate","alternative")):
            msg += "\n[조직검사 고려 포인트]\n- HRCT가 UIP/UIP-Probable이 아니거나 임상-영상 불일치 시\n- 치료 방향 결정에 중요하면 TBLC/SLB 고려\n- 가능하면 MDD 후 결정"
        self._set_text(self.output, msg)
        self.after(200, lambda: self._open_rule_and_focus(dx))

    def _open_rule_and_focus(self, dx_key: str):
        rule_name = ILD_RULE_MAP.get(dx_key, dx_key)
        try:
            open_rule_by_name(rule_name, prefer_category="ILD")
        except Exception:
            open_rule_by_name(rule_name)
        # 메인 창을 앞으로(ILD 알고리즘 창은 유지)
        try:
            self.master.lift()
            self.master.focus_force()
            self.master.attributes("-topmost", True)
            self.master.after(50, lambda: self.master.attributes("-topmost", False))
        except Exception:
            pass



    # ================= Bronchoscopy (기관지내시경) 도구 ================= #
    # - 요청 반영:
    #   2-a '정상' 선택 시 출력: "Vocal cord : normal"
    #   2-b RUL/RML/RLL 모두 정상일 때: "no endobronchial lesion or mucosal change"
    #   2-c LUL/Lingular/LLL 모두 정상일 때: "no endobronchial lesion or mucosal change"

class BronchoscopyWindow(tk.Toplevel):
    def __init__(self, master):
        super().__init__(master)
        self.title("기관지내시경")
        self.geometry("980x760")
        self.minsize(900, 700)

        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True, padx=10, pady=10)

        self.nb = nb

        self.tab_sed = ttk.Frame(nb)
        self.tab_find = ttk.Frame(nb)
        self.tab_proc = ttk.Frame(nb)

        nb.add(self.tab_sed, text="1) 진정/마취")
        nb.add(self.tab_find, text="2) 내시경 소견")
        nb.add(self.tab_proc, text="3-4) 시술/합병증")
        # Ctrl+1/2/3 for tab switching
        self._bind_ctrl_tab_shortcuts(nb, 3)

        # ---------- variables ----------
        # sedation
        self.sed_none = tk.BooleanVar(value=False)
        self.sed_mida = tk.BooleanVar(value=False)
        self.sed_mida_dose = tk.StringVar(value="")
        self.sed_fent = tk.BooleanVar(value=False)
        self.sed_fent_dose = tk.StringVar(value="")
        self.sed_prop = tk.BooleanVar(value=False)
        self.sed_prop_dose = tk.StringVar(value="")
        self.sed_etc = tk.BooleanVar(value=False)
        self.sed_etc_text = tk.StringVar(value="")

        # 2-a upper airway / vocal cord
        self.vc_status = tk.StringVar(value="")  # normal/edema/erythema/mass/movement_abnl/other
        self.vc_other = tk.StringVar(value="")

        self.sg_status = tk.StringVar(value="")  # supraglottic/subglottic normal/secretion/stenosis/mass/other
        self.sg_other = tk.StringVar(value="")

        # 2-b right bronchi
        self.rul = self._make_lobar_vars(prefix="RUL")
        self.rml = self._make_lobar_vars(prefix="RML")
        self.rll = self._make_lobar_vars(prefix="RLL", has_terminal=True, has_segment=True, has_stenosis_site=True)

        # 2-c left bronchi
        self.lul = self._make_lobar_vars(prefix="LUL")
        self.ling = self._make_lobar_vars(prefix="Lingular", has_stenosis_site=False)
        self.lll = self._make_lobar_vars(prefix="LLL", has_terminal=True, has_segment=True, has_stenosis_site=True)

        # e secretion/bleeding overall
        self.sec_amount = tk.StringVar(value="")
        self.sec_char = tk.StringVar(value="")
        self.sec_char_etc = tk.StringVar(value="")
        self.bleed = tk.StringVar(value="")
        self.bleed_site = tk.StringVar(value="")
        self.bleed_action = tk.StringVar(value="")

        # f other findings (checkboxes + free text)
        self.f_plug = tk.BooleanVar(value=False)
        self.f_asp = tk.BooleanVar(value=False)
        self.f_ext_comp = tk.BooleanVar(value=False)
        self.f_malacia = tk.BooleanVar(value=False)
        self.f_stent = tk.StringVar(value="")  # "", good, malposition, granulation, obstruction
        self.f_free = tk.StringVar(value="")

        # 3 procedures/specimens
        self.p_washing = tk.BooleanVar(value=False)
        self.p_washing_site = tk.StringVar(value="")
        self.p_bal = tk.BooleanVar(value=False)
        self.p_bal_site = tk.StringVar(value="")

        self.p_ebb = tk.BooleanVar(value=False)
        self.p_ebb_site = tk.StringVar(value="")
        self.p_ebb_n = tk.StringVar(value="")

        self.p_rebus = tk.BooleanVar(value=False)
        self.p_rebus_site = tk.StringVar(value="")
        self.p_rebus_n = tk.StringVar(value="")

        self.p_brush = tk.BooleanVar(value=False)
        self.p_brush_site = tk.StringVar(value="")

        self.p_ebus_tbna = tk.BooleanVar(value=False)
        self.p_ebus_station = tk.StringVar(value="")
        self.p_ebus_pass = tk.StringVar(value="")

        self.p_etc_proc = tk.StringVar(value="")

        # specimen requests
        self.s_cyto = tk.BooleanVar(value=False)
        self.s_histo = tk.BooleanVar(value=False)
        self.s_culture = tk.BooleanVar(value=False)
        self.s_afb = tk.BooleanVar(value=False)
        self.s_fungal = tk.BooleanVar(value=False)
        self.s_pcr = tk.BooleanVar(value=False)
        self.s_pcr_etc = tk.StringVar(value="")

        # 4 complications
        self.c_none = tk.BooleanVar(value=False)
        self.c_hypox = tk.BooleanVar(value=False)
        self.c_bleed = tk.BooleanVar(value=False)
        self.c_arr = tk.BooleanVar(value=False)
        self.c_bp = tk.BooleanVar(value=False)
        self.c_etc = tk.BooleanVar(value=False)
        self.c_etc_text = tk.StringVar(value="")

        # build UI
        self._build_sedation()
        self._build_findings()
        self._build_procedures_complications()
        self._build_output()

        # live update
        self._bind_all_traces()
        self._refresh_output()

    
    def _bind_ctrl_tab_shortcuts(self, notebook, n_tabs: int):
        """Bind Ctrl+1..n to switch notebook tabs (only when this window is active)."""
        for i in range(min(n_tabs, 9)):
            key = str(i + 1)

            def _handler(event, idx=i):
                try:
                    # Only act if the focused widget belongs to this Toplevel
                    w = event.widget
                    if w is None or w.winfo_toplevel() != self:
                        return
                    notebook.select(idx)
                    return "break"
                except Exception:
                    return

            # Use bind_all so it works no matter which child widget has focus
            self.bind_all(f"<Control-Key-{key}>", _handler, add="+")

# ---------- helpers ----------
    def _make_lobar_vars(self, prefix: str, has_terminal: bool=False, has_segment: bool=False, has_stenosis_site: bool=True):
        return {
            "prefix": prefix,
            "normal": tk.BooleanVar(value=False),
            "secretion": tk.BooleanVar(value=False),
            "secretion_segment": tk.StringVar(value="") if has_segment else None,
            "erythema": tk.BooleanVar(value=False),
            "stenosis": tk.BooleanVar(value=False),
            "stenosis_pct": tk.StringVar(value=""),
            "stenosis_site": tk.StringVar(value="") if has_stenosis_site else None,
            "mass": tk.BooleanVar(value=False),
            "mass_site": tk.StringVar(value=""),
            "ext_comp": tk.BooleanVar(value=False),
            "terminal": tk.BooleanVar(value=False) if has_terminal else None,  # 말단 폐쇄/폐색
            "other": tk.StringVar(value=""),
        }

    def _bind_var(self, v):
        try:
            v.trace_add("write", lambda *_: self._refresh_output())
        except Exception:
            pass

    def _bind_all_traces(self):
        # walk through attributes
        for attr in dir(self):
            if attr.startswith("_"):
                continue
            v = getattr(self, attr)
            if isinstance(v, (tk.StringVar, tk.BooleanVar, tk.IntVar)):
                self._bind_var(v)

        # lobar dict vars
        for lob in [self.rul, self.rml, self.rll, self.lul, self.ling, self.lll]:
            for k, v in lob.items():
                if isinstance(v, (tk.StringVar, tk.BooleanVar, tk.IntVar)):
                    self._bind_var(v)

    @staticmethod
    def _join(parts):
        return ", ".join([p for p in parts if p])

    # ---------- UI builders ----------
    def _build_sedation(self):
        f = self.tab_sed
        f.columnconfigure(0, weight=1)

        ttk.Label(f, text="진정/마취 및 기구 국소 마취", font=("맑은 고딕", 11, "bold")).grid(row=0, column=0, sticky="w", pady=(10, 8), padx=10)

        box = ttk.LabelFrame(f, text="진정/진통")
        box.grid(row=1, column=0, sticky="ew", padx=10, pady=6)
        box.columnconfigure(1, weight=1)

        r = 0
        ttk.Checkbutton(box, text="무진정", variable=self.sed_none).grid(row=r, column=0, sticky="w", padx=8, pady=4); r += 1

        row = ttk.Frame(box); row.grid(row=r, column=0, columnspan=3, sticky="ew", padx=8, pady=4); r += 1
        ttk.Checkbutton(row, text="Midazolam", variable=self.sed_mida).pack(side="left")
        ttk.Entry(row, textvariable=self.sed_mida_dose, width=10).pack(side="left", padx=6)
        ttk.Label(row, text="mg").pack(side="left")

        row = ttk.Frame(box); row.grid(row=r, column=0, columnspan=3, sticky="ew", padx=8, pady=4); r += 1
        ttk.Checkbutton(row, text="Fentanyl", variable=self.sed_fent).pack(side="left")
        ttk.Entry(row, textvariable=self.sed_fent_dose, width=10).pack(side="left", padx=6)
        ttk.Label(row, text="mcg").pack(side="left")

        row = ttk.Frame(box); row.grid(row=r, column=0, columnspan=3, sticky="ew", padx=8, pady=4); r += 1
        ttk.Checkbutton(row, text="Propofol", variable=self.sed_prop).pack(side="left")
        ttk.Entry(row, textvariable=self.sed_prop_dose, width=10).pack(side="left", padx=6)
        ttk.Label(row, text="mg").pack(side="left")

        row = ttk.Frame(box); row.grid(row=r, column=0, columnspan=3, sticky="ew", padx=8, pady=4); r += 1
        ttk.Checkbutton(row, text="기타", variable=self.sed_etc).pack(side="left")
        ttk.Entry(row, textvariable=self.sed_etc_text).pack(side="left", fill="x", expand=True, padx=6)

        btn = ttk.Frame(f)
        btn.grid(row=2, column=0, sticky="e", padx=10, pady=(8, 0))
        ttk.Button(btn, text="닫기", command=self.destroy).pack(side="right")

    def _build_findings(self):
        f = self.tab_find
        f.columnconfigure(0, weight=1)

        # make scrollable
        holder = ttk.Frame(f)
        holder.grid(row=0, column=0, sticky="nsew")
        f.rowconfigure(0, weight=1)

        canvas = tk.Canvas(holder, highlightthickness=0)
        vbar = ttk.Scrollbar(holder, orient="vertical", command=canvas.yview)
        inner = ttk.Frame(canvas)
        inner.bind("<Configure>", lambda e, c=canvas: c.configure(scrollregion=c.bbox("all")))
        canvas.create_window((0, 0), window=inner, anchor="nw")
        canvas.configure(yscrollcommand=vbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        vbar.pack(side="right", fill="y")

        def _on_mousewheel(event, c=canvas):
            c.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind("<Enter>", lambda _e, c=canvas: c.bind_all("<MouseWheel>", _on_mousewheel))
        canvas.bind("<Leave>", lambda _e, c=canvas: c.unbind_all("<MouseWheel>"))

        r = 0
        ttk.Label(inner, text="내시경 소견 (상세)", font=("맑은 고딕", 11, "bold")).grid(row=r, column=0, sticky="w", padx=10, pady=(10, 8)); r += 1

        # 2-a
        box_a = ttk.LabelFrame(inner, text="2-a) 상부기도 / 후두")
        box_a.grid(row=r, column=0, sticky="ew", padx=10, pady=6); r += 1
        box_a.columnconfigure(1, weight=1)

        row = ttk.Frame(box_a); row.grid(row=0, column=0, sticky="ew", padx=8, pady=4)
        ttk.Label(row, text="후두/성문").pack(side="left", padx=(0, 10))
        for txt, val in [("정상", "normal"), ("부종", "edema"), ("발적", "erythema"), ("종물/폴립", "mass"), ("움직임 이상(마비/제한)", "movement_abnl"), ("기타", "other")]:
            ttk.Radiobutton(row, text=txt, variable=self.vc_status, value=val).pack(side="left", padx=6)
        row2 = ttk.Frame(box_a); row2.grid(row=1, column=0, sticky="ew", padx=8, pady=4)
        ttk.Label(row2, text="기타").pack(side="left")
        ttk.Entry(row2, textvariable=self.vc_other).pack(side="left", fill="x", expand=True, padx=6)

        row = ttk.Frame(box_a); row.grid(row=2, column=0, sticky="ew", padx=8, pady=4)
        ttk.Label(row, text="성문 상부/하부").pack(side="left", padx=(0, 10))
        for txt, val in [("정상", "normal"), ("분비물", "secretion"), ("협착", "stenosis"), ("종물", "mass"), ("기타", "other")]:
            ttk.Radiobutton(row, text=txt, variable=self.sg_status, value=val).pack(side="left", padx=6)
        row2 = ttk.Frame(box_a); row2.grid(row=3, column=0, sticky="ew", padx=8, pady=4)
        ttk.Label(row2, text="기타").pack(side="left")
        ttk.Entry(row2, textvariable=self.sg_other).pack(side="left", fill="x", expand=True, padx=6)

        # 2-b right
        box_b = ttk.LabelFrame(inner, text="2-b) 우측 기관지")
        box_b.grid(row=r, column=0, sticky="ew", padx=10, pady=6); r += 1
        self._build_lobar_block(box_b, 0, self.rul, label="RUL (우상엽 bronchus)")
        self._build_lobar_block(box_b, 1, self.rml, label="RML (우중엽 bronchus)")
        self._build_lobar_block(box_b, 2, self.rll, label="RLL (우하엽 bronchus)")

        # 2-c left
        box_c = ttk.LabelFrame(inner, text="2-c) 좌측 기관지")
        box_c.grid(row=r, column=0, sticky="ew", padx=10, pady=6); r += 1
        self._build_lobar_block(box_c, 0, self.lul, label="LUL")
        self._build_lobar_block(box_c, 1, self.ling, label="Lingular segment")
        self._build_lobar_block(box_c, 2, self.lll, label="LLL")

        # e/f
        box_e = ttk.LabelFrame(inner, text="e) 분비물/출혈 (전반 평가)")
        box_e.grid(row=r, column=0, sticky="ew", padx=10, pady=6); r += 1
        row = ttk.Frame(box_e); row.pack(fill="x", padx=8, pady=4)
        ttk.Label(row, text="분비물 양").pack(side="left", padx=(0, 10))
        for txt in ["거의 없음", "적음", "중등도", "많음"]:
            ttk.Radiobutton(row, text=txt, variable=self.sec_amount, value=txt).pack(side="left", padx=6)

        row = ttk.Frame(box_e); row.pack(fill="x", padx=8, pady=4)
        ttk.Label(row, text="분비물 성상").pack(side="left", padx=(0, 10))
        for txt in ["맑음", "점액성", "화농성", "점액농성", "혈성", "기타"]:
            ttk.Radiobutton(row, text=txt, variable=self.sec_char, value=txt).pack(side="left", padx=6)
        row2 = ttk.Frame(box_e); row2.pack(fill="x", padx=8, pady=(0, 6))
        ttk.Label(row2, text="기타").pack(side="left")
        ttk.Entry(row2, textvariable=self.sec_char_etc).pack(side="left", fill="x", expand=True, padx=6)

        row = ttk.Frame(box_e); row.pack(fill="x", padx=8, pady=4)
        ttk.Label(row, text="출혈").pack(side="left", padx=(0, 10))
        for txt, val in [("없음", "none"), ("소량 oozing", "oozing"), ("국소 출혈", "focal"), ("활동성 출혈", "active")]:
            ttk.Radiobutton(row, text=txt, variable=self.bleed, value=val).pack(side="left", padx=6)
        row2 = ttk.Frame(box_e); row2.pack(fill="x", padx=8, pady=(0, 6))
        ttk.Label(row2, text="부위/조치").pack(side="left")
        ttk.Entry(row2, textvariable=self.bleed_site, width=24).pack(side="left", padx=6)
        ttk.Entry(row2, textvariable=self.bleed_action).pack(side="left", fill="x", expand=True, padx=6)

        box_f = ttk.LabelFrame(inner, text="f) 기타 특이 소견")
        box_f.grid(row=r, column=0, sticky="ew", padx=10, pady=6); r += 1
        row = ttk.Frame(box_f); row.pack(fill="x", padx=8, pady=4)
        ttk.Checkbutton(row, text="기도 내 plug/점액전", variable=self.f_plug).pack(side="left", padx=6)
        ttk.Checkbutton(row, text="흡인 의심 물질", variable=self.f_asp).pack(side="left", padx=6)
        ttk.Checkbutton(row, text="기도 외부압박 소견", variable=self.f_ext_comp).pack(side="left", padx=6)
        ttk.Checkbutton(row, text="기도 연화증/dynamic collapse 의심", variable=self.f_malacia).pack(side="left", padx=6)

        row2 = ttk.Frame(box_f); row2.pack(fill="x", padx=8, pady=4)
        ttk.Label(row2, text="스텐트 상태").pack(side="left", padx=(0, 10))
        for txt, val in [("양호", "good"), ("malposition", "malposition"), ("과립조직", "granulation"), ("폐쇄/막힘", "obstruction")]:
            ttk.Radiobutton(row2, text=txt, variable=self.f_stent, value=val).pack(side="left", padx=6)

        row3 = ttk.Frame(box_f); row3.pack(fill="x", padx=8, pady=(0, 8))
        ttk.Label(row3, text="자유기입(종합 인상/특이 소견)").pack(anchor="w")
        ttk.Entry(row3, textvariable=self.f_free).pack(fill="x", expand=True, pady=(2, 0))

        inner.columnconfigure(0, weight=1)

    def _build_lobar_block(self, parent, row_idx: int, lobar: dict, label: str):
        block = ttk.LabelFrame(parent, text=label)
        block.grid(row=row_idx, column=0, sticky="ew", padx=8, pady=6)
        block.columnconfigure(0, weight=1)

        top = ttk.Frame(block); top.pack(fill="x", padx=6, pady=4)
        ttk.Checkbutton(top, text="정상", variable=lobar["normal"]).pack(side="left", padx=6)
        ttk.Checkbutton(top, text="분비물 증가", variable=lobar["secretion"]).pack(side="left", padx=6)
        if lobar.get("secretion_segment") is not None:
            ttk.Label(top, text="segment").pack(side="left", padx=(10, 2))
            ttk.Entry(top, textvariable=lobar["secretion_segment"], width=14).pack(side="left")

        mid = ttk.Frame(block); mid.pack(fill="x", padx=6, pady=4)
        ttk.Checkbutton(mid, text="점막 발적/부종", variable=lobar["erythema"]).pack(side="left", padx=6)
        ttk.Checkbutton(mid, text="내강 협착", variable=lobar["stenosis"]).pack(side="left", padx=6)
        ttk.Label(mid, text="약").pack(side="left")
        ttk.Entry(mid, textvariable=lobar["stenosis_pct"], width=6).pack(side="left", padx=4)
        ttk.Label(mid, text="%").pack(side="left")
        if lobar.get("stenosis_site") is not None:
            ttk.Label(mid, text="부위").pack(side="left", padx=(10, 2))
            ttk.Entry(mid, textvariable=lobar["stenosis_site"], width=18).pack(side="left")

        low = ttk.Frame(block); low.pack(fill="x", padx=6, pady=4)
        ttk.Checkbutton(low, text="종물/결절", variable=lobar["mass"]).pack(side="left", padx=6)
        ttk.Label(low, text="위치").pack(side="left", padx=(10, 2))
        ttk.Entry(low, textvariable=lobar["mass_site"], width=24).pack(side="left")

        low2 = ttk.Frame(block); low2.pack(fill="x", padx=6, pady=4)
        ttk.Checkbutton(low2, text="외부압박 의심", variable=lobar["ext_comp"]).pack(side="left", padx=6)
        if lobar.get("terminal") is not None:
            ttk.Checkbutton(low2, text="말단 폐쇄/폐색 소견", variable=lobar["terminal"]).pack(side="left", padx=6)
        ttk.Label(low2, text="기타").pack(side="left", padx=(10, 2))
        ttk.Entry(low2, textvariable=lobar["other"]).pack(side="left", fill="x", expand=True)

        # if 정상 체크하면 나머지 자동 해제(편의)
        def _on_normal(*_):
            if lobar["normal"].get():
                for k in ["secretion", "erythema", "stenosis", "mass", "ext_comp"]:
                    try:
                        lobar[k].set(False)
                    except Exception:
                        pass
                if lobar.get("terminal") is not None:
                    lobar["terminal"].set(False)
        try:
            lobar["normal"].trace_add("write", _on_normal)
        except Exception:
            pass

    def _build_procedures_complications(self):
        f = self.tab_proc
        f.columnconfigure(0, weight=1)

        holder = ttk.Frame(f)
        holder.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        f.rowconfigure(0, weight=1)

        # procedures
        box_p = ttk.LabelFrame(holder, text="3) 시행 시술 및 검체")
        box_p.pack(fill="x", pady=6)

        row = ttk.Frame(box_p); row.pack(fill="x", padx=8, pady=4)
        ttk.Checkbutton(row, text="Bronchial washing", variable=self.p_washing).pack(side="left", padx=6)
        ttk.Label(row, text="부위").pack(side="left")
        ttk.Entry(row, textvariable=self.p_washing_site, width=24).pack(side="left", padx=6)

        row = ttk.Frame(box_p); row.pack(fill="x", padx=8, pady=4)
        ttk.Checkbutton(row, text="BAL", variable=self.p_bal).pack(side="left", padx=6)
        ttk.Label(row, text="부위").pack(side="left")
        ttk.Entry(row, textvariable=self.p_bal_site, width=24).pack(side="left", padx=6)

        row = ttk.Frame(box_p); row.pack(fill="x", padx=8, pady=4)
        ttk.Checkbutton(row, text="Endobronchial biopsy", variable=self.p_ebb).pack(side="left", padx=6)
        ttk.Label(row, text="부위").pack(side="left")
        ttk.Entry(row, textvariable=self.p_ebb_site, width=24).pack(side="left", padx=6)
        ttk.Label(row, text="개수").pack(side="left", padx=(10, 2))
        ttk.Entry(row, textvariable=self.p_ebb_n, width=6).pack(side="left")

        row = ttk.Frame(box_p); row.pack(fill="x", padx=8, pady=4)
        ttk.Checkbutton(row, text="Radial EBUS", variable=self.p_rebus).pack(side="left", padx=6)
        ttk.Label(row, text="부위").pack(side="left")
        ttk.Entry(row, textvariable=self.p_rebus_site, width=24).pack(side="left", padx=6)
        ttk.Label(row, text="개수").pack(side="left", padx=(10, 2))
        ttk.Entry(row, textvariable=self.p_rebus_n, width=6).pack(side="left")

        row = ttk.Frame(box_p); row.pack(fill="x", padx=8, pady=4)
        ttk.Checkbutton(row, text="Brushing", variable=self.p_brush).pack(side="left", padx=6)
        ttk.Label(row, text="부위").pack(side="left")
        ttk.Entry(row, textvariable=self.p_brush_site, width=24).pack(side="left", padx=6)

        row = ttk.Frame(box_p); row.pack(fill="x", padx=8, pady=4)
        ttk.Checkbutton(row, text="EBUS-TBNA", variable=self.p_ebus_tbna).pack(side="left", padx=6)
        ttk.Label(row, text="station").pack(side="left")
        ttk.Entry(row, textvariable=self.p_ebus_station, width=12).pack(side="left", padx=6)
        ttk.Label(row, text="pass").pack(side="left")
        ttk.Entry(row, textvariable=self.p_ebus_pass, width=6).pack(side="left", padx=6)

        row = ttk.Frame(box_p); row.pack(fill="x", padx=8, pady=4)
        ttk.Label(row, text="기타 시술").pack(side="left", padx=(0, 10))
        ttk.Entry(row, textvariable=self.p_etc_proc).pack(side="left", fill="x", expand=True)

        # specimen requests
        box_s = ttk.LabelFrame(holder, text="검체 의뢰")
        box_s.pack(fill="x", pady=6)
        row = ttk.Frame(box_s); row.pack(fill="x", padx=8, pady=4)
        ttk.Checkbutton(row, text="세포병리", variable=self.s_cyto).pack(side="left", padx=6)
        ttk.Checkbutton(row, text="조직병리", variable=self.s_histo).pack(side="left", padx=6)
        ttk.Checkbutton(row, text="세균 배양", variable=self.s_culture).pack(side="left", padx=6)
        ttk.Checkbutton(row, text="AFB/결핵", variable=self.s_afb).pack(side="left", padx=6)
        ttk.Checkbutton(row, text="진균", variable=self.s_fungal).pack(side="left", padx=6)
        ttk.Checkbutton(row, text="PCR/기타", variable=self.s_pcr).pack(side="left", padx=6)
        row2 = ttk.Frame(box_s); row2.pack(fill="x", padx=8, pady=(0, 6))
        ttk.Label(row2, text="PCR/기타").pack(side="left")
        ttk.Entry(row2, textvariable=self.s_pcr_etc).pack(side="left", fill="x", expand=True, padx=6)

        # complications
        box_c = ttk.LabelFrame(holder, text="4) 합병증 및 시술 중 문제")
        box_c.pack(fill="x", pady=6)
        row = ttk.Frame(box_c); row.pack(fill="x", padx=8, pady=4)
        ttk.Checkbutton(row, text="없음", variable=self.c_none).pack(side="left", padx=6)
        ttk.Checkbutton(row, text="저산소혈증", variable=self.c_hypox).pack(side="left", padx=6)
        ttk.Checkbutton(row, text="출혈", variable=self.c_bleed).pack(side="left", padx=6)
        ttk.Checkbutton(row, text="부정맥", variable=self.c_arr).pack(side="left", padx=6)
        ttk.Checkbutton(row, text="저혈압/고혈압", variable=self.c_bp).pack(side="left", padx=6)
        ttk.Checkbutton(row, text="기타", variable=self.c_etc).pack(side="left", padx=6)
        row2 = ttk.Frame(box_c); row2.pack(fill="x", padx=8, pady=(0, 6))
        ttk.Label(row2, text="기타").pack(side="left")
        ttk.Entry(row2, textvariable=self.c_etc_text).pack(side="left", fill="x", expand=True, padx=6)

        def _on_none(*_):
            if self.c_none.get():
                for v in [self.c_hypox, self.c_bleed, self.c_arr, self.c_bp, self.c_etc]:
                    v.set(False)
        try:
            self.c_none.trace_add("write", _on_none)
        except Exception:
            pass

    def _build_output(self):
        # Bottom, persistent output area (no separate '출력' tab)
        f = ttk.Labelframe(self, text="출력")
        f.pack(fill="both", expand=False, padx=10, pady=(0, 10))
        f.columnconfigure(0, weight=1)
        f.rowconfigure(0, weight=1)

        # height keeps it compact while always visible
        self.output = scrolledtext.ScrolledText(f, wrap="word", height=10, font=("맑은 고딕", 10))
        self.output.grid(row=0, column=0, sticky="nsew", padx=10, pady=(10, 6))

        btn = ttk.Frame(f)
        btn.grid(row=1, column=0, sticky="e", padx=10, pady=(0, 10))
        ttk.Button(btn, text="복사", command=self._copy).pack(side="left")
        ttk.Button(btn, text="닫기", command=self.destroy).pack(side="left", padx=(10, 0))


        self.output = scrolledtext.ScrolledText(f, wrap="word", font=("맑은 고딕", 10))
        self.output.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)

        btn = ttk.Frame(f)
        btn.grid(row=1, column=0, sticky="e", padx=10, pady=(0, 10))
        ttk.Button(btn, text="복사", command=self._copy).pack(side="left")
        ttk.Button(btn, text="닫기", command=self.destroy).pack(side="left", padx=(10, 0))

    def _copy(self):
        text_out = self.output.get("1.0", "end-1c")
        try:
            self.clipboard_clear()
            self.clipboard_append(text_out)
            messagebox.showinfo("복사됨", "기관지내시경 결과를 클립보드에 복사했습니다.")
        except Exception:
            pass

    # ---------- output compose ----------
    def _sedation_text(self):
        parts = []
        if self.sed_none.get():
            parts.append("무진정")
        if self.sed_mida.get():
            dose = (self.sed_mida_dose.get() or "").strip()
            parts.append(f"Midazolam {dose} mg".strip())
        if self.sed_fent.get():
            dose = (self.sed_fent_dose.get() or "").strip()
            parts.append(f"Fentanyl {dose} mcg".strip())
        if self.sed_prop.get():
            dose = (self.sed_prop_dose.get() or "").strip()
            parts.append(f"Propofol {dose} mg".strip())
        if self.sed_etc.get():
            etc = (self.sed_etc_text.get() or "").strip()
            if etc:
                parts.append(f"기타: {etc}")
            else:
                parts.append("기타")
        return self._join(parts)

    def _upper_airway_text(self):
        lines = []
        vc = (self.vc_status.get() or "").strip()
        if vc == "normal":
            lines.append("Vocal cord : normal")
        elif vc:
            mapping = {
                "edema": "Vocal cord : edema",
                "erythema": "Vocal cord : erythema",
                "mass": "Vocal cord : mass/polyp",
                "movement_abnl": "Vocal cord : movement abnormality",
                "other": "",
            }
            if vc == "other":
                other = (self.vc_other.get() or "").strip()
                if other:
                    lines.append(f"Vocal cord : {other}")
            else:
                lines.append(mapping.get(vc, f"Vocal cord : {vc}"))

        sg = (self.sg_status.get() or "").strip()
        if sg:
            if sg == "normal":
                lines.append("Supraglottic/Subglottic : normal")
            elif sg == "other":
                other = (self.sg_other.get() or "").strip()
                if other:
                    lines.append(f"Supraglottic/Subglottic : {other}")
            else:
                lines.append(f"Supraglottic/Subglottic : {sg}")
        return lines

    def _lobar_all_normal(self, *lobars):
        return all(l["normal"].get() for l in lobars)

    def _lobar_text(self, lobar: dict):
        if lobar["normal"].get():
            return f"{lobar['prefix']}: normal"
        parts = []
        if lobar["secretion"].get():
            seg = (lobar.get("secretion_segment").get() if lobar.get("secretion_segment") is not None else "")
            seg = (seg or "").strip()
            parts.append("increased secretion" + (f" (segment: {seg})" if seg else ""))
        if lobar["erythema"].get():
            parts.append("mucosal erythema/edema")
        if lobar["stenosis"].get():
            pct = (lobar["stenosis_pct"].get() or "").strip()
            site = ""
            if lobar.get("stenosis_site") is not None:
                site = (lobar["stenosis_site"].get() or "").strip()
            msg = "luminal narrowing"
            if pct:
                msg += f" (~{pct}%)"
            if site:
                msg += f" (site: {site})"
            parts.append(msg)
        if lobar["mass"].get():
            site = (lobar["mass_site"].get() or "").strip()
            parts.append("mass/nodule" + (f" (site: {site})" if site else ""))
        if lobar["ext_comp"].get():
            parts.append("suspected extrinsic compression")
        if lobar.get("terminal") is not None and lobar["terminal"].get():
            parts.append("distal occlusion/obstruction")
        other = (lobar["other"].get() or "").strip()
        if other:
            parts.append(other)
        if not parts:
            # 아무것도 선택 안 된 경우
            return f"{lobar['prefix']}: (no selection)"
        return f"{lobar['prefix']}: " + "; ".join(parts)

    def _secretion_bleeding_text(self):
        parts = []
        if (self.sec_amount.get() or "").strip():
            parts.append(f"Secretion amount: {self.sec_amount.get()}")
        sc = (self.sec_char.get() or "").strip()
        if sc:
            if sc == "기타":
                etc = (self.sec_char_etc.get() or "").strip()
                parts.append("Secretion character: " + (etc if etc else "etc"))
            else:
                parts.append(f"Secretion character: {sc}")
        b = (self.bleed.get() or "").strip()
        if b:
            mapping = {"none": "Bleeding: none", "oozing": "Bleeding: minor oozing", "focal": "Bleeding: focal", "active": "Bleeding: active"}
            line = mapping.get(b, f"Bleeding: {b}")
            site = (self.bleed_site.get() or "").strip()
            action = (self.bleed_action.get() or "").strip()
            if site:
                line += f" (site: {site})"
            if action:
                line += f" (action: {action})"
            parts.append(line)
        return parts

    def _other_findings_text(self):
        parts = []
        if self.f_plug.get():
            parts.append("airway plug/mucus plug")
        if self.f_asp.get():
            parts.append("suspected aspiration material")
        if self.f_ext_comp.get():
            parts.append("extrinsic compression")
        if self.f_malacia.get():
            parts.append("suspected tracheobronchomalacia/dynamic collapse")
        st = (self.f_stent.get() or "").strip()
        if st:
            parts.append(f"stent: {st}")
        free = (self.f_free.get() or "").strip()
        if free:
            parts.append(free)
        return parts

    def _procedures_text(self):
        parts = []
        if self.p_washing.get():
            site = (self.p_washing_site.get() or "").strip()
            parts.append("Bronchial washing" + (f" (site: {site})" if site else ""))
        if self.p_bal.get():
            site = (self.p_bal_site.get() or "").strip()
            parts.append("BAL" + (f" (site: {site})" if site else ""))
        if self.p_ebb.get():
            site = (self.p_ebb_site.get() or "").strip()
            n = (self.p_ebb_n.get() or "").strip()
            extra = []
            if site: extra.append(f"site: {site}")
            if n: extra.append(f"n={n}")
            parts.append("Endobronchial biopsy" + (f" ({', '.join(extra)})" if extra else ""))
        if self.p_rebus.get():
            site = (self.p_rebus_site.get() or "").strip()
            n = (self.p_rebus_n.get() or "").strip()
            extra = []
            if site: extra.append(f"site: {site}")
            if n: extra.append(f"n={n}")
            parts.append("Radial EBUS" + (f" ({', '.join(extra)})" if extra else ""))
        if self.p_brush.get():
            site = (self.p_brush_site.get() or "").strip()
            parts.append("Brushing" + (f" (site: {site})" if site else ""))
        if self.p_ebus_tbna.get():
            st = (self.p_ebus_station.get() or "").strip()
            ps = (self.p_ebus_pass.get() or "").strip()
            extra = []
            if st: extra.append(f"station: {st}")
            if ps: extra.append(f"pass: {ps}")
            parts.append("EBUS-TBNA" + (f" ({', '.join(extra)})" if extra else ""))
        etc = (self.p_etc_proc.get() or "").strip()
        if etc:
            parts.append(f"Other procedure: {etc}")
        return parts

    def _specimen_requests_text(self):
        req = []
        if self.s_cyto.get(): req.append("Cytology")
        if self.s_histo.get(): req.append("Histology")
        if self.s_culture.get(): req.append("Bacterial culture")
        if self.s_afb.get(): req.append("AFB/TB")
        if self.s_fungal.get(): req.append("Fungal")
        if self.s_pcr.get():
            etc = (self.s_pcr_etc.get() or "").strip()
            req.append("PCR/Other" + (f": {etc}" if etc else ""))
        return req

    def _complications_text(self):
        parts = []
        if self.c_none.get():
            return ["none"]
        if self.c_hypox.get(): parts.append("hypoxemia")
        if self.c_bleed.get(): parts.append("bleeding")
        if self.c_arr.get(): parts.append("arrhythmia")
        if self.c_bp.get(): parts.append("hypotension/hypertension")
        if self.c_etc.get():
            etc = (self.c_etc_text.get() or "").strip()
            parts.append(etc if etc else "etc")
        return parts

    def _compose(self) -> str:
        lines = []
        sed = self._sedation_text()
        if sed:
            lines.append("1. Sedation/Analgesia: " + sed)

        lines.append("")
        lines.append("2. Findings:")

        # 2-a
        ua = self._upper_airway_text()
        if ua:
            lines.append("  a) Upper airway / Larynx")
            for x in ua:
                lines.append("     - " + x)

        # 2-b right
        lines.append("  b) Right bronchus")
        if self._lobar_all_normal(self.rul, self.rml, self.rll):
            lines.append("     - no endobronchial lesion or mucosal change")
        else:
            for lob in [self.rul, self.rml, self.rll]:
                lines.append("     - " + self._lobar_text(lob))

        # 2-c left
        lines.append("  c) Left bronchus")
        if self._lobar_all_normal(self.lul, self.ling, self.lll):
            lines.append("     - no endobronchial lesion or mucosal change")
        else:
            for lob in [self.lul, self.ling, self.lll]:
                lines.append("     - " + self._lobar_text(lob))

        # e
        se = self._secretion_bleeding_text()
        if se:
            lines.append("  e) Secretion/Bleeding (overall)")
            for x in se:
                lines.append("     - " + x)

        # f
        oth = self._other_findings_text()
        if oth:
            lines.append("  f) Other findings")
            lines.append("     - " + "; ".join(oth))

        # procedures/specimens
        proc = self._procedures_text()
        if proc:
            lines.append("")
            lines.append("3. Procedures:")
            for x in proc:
                lines.append("  - " + x)

        req = self._specimen_requests_text()
        if req:
            lines.append("  Specimen requests: " + ", ".join(req))

        comp = self._complications_text()
        if comp:
            lines.append("")
            lines.append("4. Complications:")
            lines.append("  - " + ", ".join(comp))

        return "\n".join(lines).strip() + "\n"

    def _refresh_output(self):
        try:
            text_out = self._compose()
            self.output.delete("1.0", "end")
            self.output.insert("1.0", text_out)
        except Exception:
            pass


def open_bronchoscopy(master):
    try:
        BronchoscopyWindow(master)
    except Exception as e:
        messagebox.showerror("기관지내시경 오류", f"기관지내시경 창 실행 중 오류가 발생했습니다:\n{e}")



# ================= Operability (폐수술 전 평가) 도구 ================= #
# - ppoFEV1/ppoDLCO 계산(분절 기반 간이)
# - ARISCAT 점수 계산

class OperabilityWindow(tk.Toplevel):
    def __init__(self, master):
        super().__init__(master)
        self.title("Operability (폐수술 전 평가)")
        self.geometry("900x650")

        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True)

        self.tab_ppo = ttk.Frame(nb)
        self.tab_ariscat = ttk.Frame(nb)
        nb.add(self.tab_ppo, text="1) ppoFEV1/ppoDLCO")
        nb.add(self.tab_ariscat, text="2) ARISCAT")

        self._build_ppo()
        self._build_ariscat()

    # ---------- ppo ----------
    def _build_ppo(self):
        f = self.tab_ppo
        f.columnconfigure(1, weight=1)

        ttk.Label(f, text="ppoFEV1 / ppoDLCO 계산 (분절 기반 간이)", font=("맑은 고딕", 11, "bold")).grid(
            row=0, column=0, columnspan=3, sticky="w", padx=10, pady=(10, 6)
        )

        note = (
            "간이 공식(분절 기반):\n"
            "  ppo%pred = preop%pred × (1 - 제거분절수 / 총분절수)\n"
            "총분절수는 기본 19(양측)로 설정했습니다.\n"
            "* 실제 임상에서는 관류/환기 스캔 기반 계산 등 기관 프로토콜을 우선하세요."
        )
        ttk.Label(f, text=note, foreground="#444444").grid(row=1, column=0, columnspan=3, sticky="w", padx=10, pady=(0, 10))

        self.preop_fev1 = tk.StringVar(value="")
        self.preop_dlco = tk.StringVar(value="")
        self.total_segments = tk.StringVar(value="19")
        self.resect_segments = tk.StringVar(value="")

        def _row(r, label, var):
            ttk.Label(f, text=label).grid(row=r, column=0, sticky="w", padx=10, pady=4)
            e = ttk.Entry(f, textvariable=var, width=12)
            e.grid(row=r, column=1, sticky="w", pady=4)
            return e

        _row(2, "Preop FEV1 (%pred)", self.preop_fev1)
        _row(3, "Preop DLCO (%pred)", self.preop_dlco)
        _row(4, "총 폐 분절수 (기본 19)", self.total_segments)
        _row(5, "제거(절제) 분절수", self.resect_segments)

        
        # 폐엽 선택(버튼/복수 선택 가능) → 자동 분절수 합산
        #  - 우상엽(RUL)=3, 우중엽(RML)=2, 우하엽(RLL)=5, 좌상엽(LUL)=5, 좌하엽(LLL)=4 (총 19)
        #  - bilobectomy/전절제 등 복수 선택 시 자동 합산됩니다.
        self.lobe_vars = {
            "RUL(우상엽)": tk.IntVar(value=0),
            "RML(우중엽)": tk.IntVar(value=0),
            "RLL(우하엽)": tk.IntVar(value=0),
            "LUL(좌상엽)": tk.IntVar(value=0),
            "LLL(좌하엽)": tk.IntVar(value=0),
        }
        self.lobe_seg_map = {
            "RUL(우상엽)": 3,
            "RML(우중엽)": 2,
            "RLL(우하엽)": 5,
            "LUL(좌상엽)": 5,
            "LLL(좌하엽)": 4,
        }

        ttk.Label(f, text="폐엽 선택(복수 선택 가능)").grid(row=6, column=0, sticky="w", padx=10, pady=(8, 2))
        lobe_frame = ttk.Frame(f)
        lobe_frame.grid(row=6, column=1, columnspan=2, sticky="w", padx=10, pady=(8, 2))

        def _update_resect_from_lobes(*_):
            total = 0
            for k, var in self.lobe_vars.items():
                if int(var.get()) == 1:
                    total += int(self.lobe_seg_map.get(k, 0))
            # 선택이 없으면 기존 입력 유지(수동 입력 허용)
            if total > 0:
                self.resect_segments.set(str(total))

        # 버튼형 토글(복수 선택 가능)
        col = 0
        for name in ["RUL(우상엽)", "RML(우중엽)", "RLL(우하엽)", "LUL(좌상엽)", "LLL(좌하엽)"]:
            cb = tk.Checkbutton(
                lobe_frame,
                text=name,
                variable=self.lobe_vars[name],
                indicatoron=False,   # 버튼처럼 보이게
                command=_update_resect_from_lobes,
                padx=10,
                pady=4
            )
            cb.grid(row=0, column=col, sticky="w", padx=(0, 10))
            col += 1

        self.lbl_ppo = ttk.Label(f, text="ppoFEV1: -    ppoDLCO: -", font=("맑은 고딕", 11, "bold"), foreground="blue")
        self.lbl_ppo.grid(row=7, column=0, columnspan=3, sticky="w", padx=10, pady=(12, 4))

        self.txt_ppo_detail = tk.Text(f, height=8, wrap="word")
        self.txt_ppo_detail.grid(row=8, column=0, columnspan=3, sticky="nsew", padx=10, pady=(6, 10))
        f.rowconfigure(8, weight=1)

        def _calc(*_):
            try:
                def _opt_float(s: str):
                    s = (s or "").strip()
                    if s == "":
                        return None
                    return float(s)

                fev1 = _opt_float(self.preop_fev1.get())
                dlco = _opt_float(self.preop_dlco.get())
                total = float((self.total_segments.get() or "19").strip() or "19")
                # 제거 분절수는 비워두면 0으로 간주(=절제 없음)
                resect = _opt_float(self.resect_segments.get())
                if resect is None:
                    resect = 0.0


                if total <= 0:
                    raise ValueError("총 분절수는 0보다 커야 합니다.")
                frac = max(0.0, min(1.0, 1.0 - (resect / total)))
                ppo_fev1 = (fev1 * frac) if (fev1 is not None) else None
                ppo_dlco = (dlco * frac) if (dlco is not None) else None

                # 부분 입력 허용: FEV1만 입력하면 ppoFEV1만, DLCO만 입력하면 ppoDLCO만 계산
                fev1_txt = f"{ppo_fev1:.1f}%pred" if ppo_fev1 is not None else "-"
                dlco_txt = f"{ppo_dlco:.1f}%pred" if ppo_dlco is not None else "-"
                self.lbl_ppo.config(text=f"ppoFEV1: {fev1_txt}    ppoDLCO: {dlco_txt}")

                detail_lines = []
                detail_lines.append("계산:")
                detail_lines.append(f"  remaining fraction = 1 - (resect/total) = 1 - ({resect:g}/{total:g}) = {frac:.3f}")

                if ppo_fev1 is not None:
                    detail_lines.append(f"  ppoFEV1 = {fev1:g} × {frac:.3f} = {ppo_fev1:.1f} (%pred)")
                else:
                    detail_lines.append("  ppoFEV1 = (Preop FEV1 입력 시 계산)")

                if ppo_dlco is not None:
                    detail_lines.append(f"  ppoDLCO = {dlco:g} × {frac:.3f} = {ppo_dlco:.1f} (%pred)")
                else:
                    detail_lines.append("  ppoDLCO = (Preop DLCO 입력 시 계산)")

                detail = "\n".join(detail_lines)
                self._set_text(self.txt_ppo_detail, detail)
            except Exception as e:
                self.lbl_ppo.config(text="ppoFEV1: -    ppoDLCO: -")
                self._set_text(
                    self.txt_ppo_detail,
                    f"입력값을 확인해 주세요.\n- 숫자만 입력 (예: 65, 42.5)\n- 제거 분절수는 비워두면 0(절제 없음)으로 계산됩니다.\n\n오류: {e}"
                )

        # 실시간 계산
        for v in [self.preop_fev1, self.preop_dlco, self.total_segments, self.resect_segments]:
            v.trace_add("write", _calc)

        btn = ttk.Frame(f)
        btn.grid(row=8, column=0, columnspan=3, sticky="e", padx=10, pady=(0, 10))
        ttk.Button(btn, text="닫기", command=self.destroy).pack(side="right")

    def _set_text(self, widget: tk.Text, s: str):
        widget.config(state="normal")
        widget.delete("1.0", tk.END)
        widget.insert(tk.END, s)
        widget.config(state="disabled")

    # ---------- ARISCAT ----------
    def _build_ariscat(self):
        f = self.tab_ariscat
        f.columnconfigure(1, weight=1)

        ttk.Label(f, text="ARISCAT (Postoperative pulmonary complications) 점수", font=("맑은 고딕", 11, "bold")).grid(
            row=0, column=0, columnspan=3, sticky="w", padx=10, pady=(10, 6)
        )

        self.age = tk.StringVar(value="")
        self.spo2 = tk.StringVar(value="")
        self.recent_inf = tk.StringVar(value="아니오")
        self.hb_low = tk.StringVar(value="아니오")
        self.incision = tk.StringVar(value="말초")
        self.duration = tk.StringVar(value="")
        self.emerg = tk.StringVar(value="아니오")

        def _row_entry(r, label, var, hint=""):
            ttk.Label(f, text=label).grid(row=r, column=0, sticky="w", padx=10, pady=4)
            e = ttk.Entry(f, textvariable=var, width=12)
            e.grid(row=r, column=1, sticky="w", pady=4)
            if hint:
                ttk.Label(f, text=hint, foreground="#444444").grid(row=r, column=2, sticky="w", padx=(6,0))
            return e

        _row_entry(1, "나이 (세)", self.age)
        _row_entry(2, "수술 전 SpO₂ (%)", self.spo2, "실온 공기 기준 권장")
        ttk.Label(f, text="최근 1개월 호흡기 감염",).grid(row=3, column=0, sticky="w", padx=10, pady=4)
        ttk.Combobox(f, textvariable=self.recent_inf, values=["아니오", "예"], state="readonly", width=10).grid(row=3, column=1, sticky="w", pady=4)

        ttk.Label(f, text="빈혈 (Hb < 10 g/dL)").grid(row=4, column=0, sticky="w", padx=10, pady=4)
        ttk.Combobox(f, textvariable=self.hb_low, values=["아니오", "예"], state="readonly", width=10).grid(row=4, column=1, sticky="w", pady=4)

        ttk.Label(f, text="수술 절개/부위").grid(row=5, column=0, sticky="w", padx=10, pady=4)
        ttk.Combobox(
            f,
            textvariable=self.incision,
            values=["말초", "상복부", "흉부(흉강내)"],
            state="readonly",
            width=12,
        ).grid(row=5, column=1, sticky="w", pady=4)

        _row_entry(6, "수술 시간 (분)", self.duration)
        ttk.Label(f, text="응급 수술").grid(row=7, column=0, sticky="w", padx=10, pady=4)
        ttk.Combobox(f, textvariable=self.emerg, values=["아니오", "예"], state="readonly", width=10).grid(row=7, column=1, sticky="w", pady=4)

        self.lbl_ariscat = ttk.Label(f, text="ARISCAT 점수: -   위험도: -", font=("맑은 고딕", 11, "bold"), foreground="darkred")
        self.lbl_ariscat.grid(row=8, column=0, columnspan=3, sticky="w", padx=10, pady=(12, 4))

        self.txt_ariscat_detail = tk.Text(f, height=10, wrap="word")
        self.txt_ariscat_detail.grid(row=9, column=0, columnspan=3, sticky="nsew", padx=10, pady=(6, 10))
        f.rowconfigure(9, weight=1)

        def _score():
            # ARISCAT scoring (Canet et al. 2010) commonly used cutoffs:
            # Age: 51-80=3, >80=16
            # SpO2: >=96=0, 91-95=8, <=90=24
            # Recent respiratory infection (last month): 17
            # Preop anemia (Hb <10): 11
            # Surgical incision: peripheral=0, upper abdominal=15, intrathoracic=24
            # Duration: <2h=0, 2-3h=16, >3h=23  (minutes <120, 120-180, >180)
            # Emergency: 8
            pts = 0
            parts = []

            # Age
            try:
                age = float((self.age.get() or "").strip())
            except Exception:
                age = None
            if age is not None:
                if age > 80:
                    pts += 16; parts.append("나이 >80: +16")
                elif age >= 51:
                    pts += 3; parts.append("나이 51–80: +3")
                else:
                    parts.append("나이 ≤50: +0")

            # SpO2
            try:
                sp = float((self.spo2.get() or "").strip())
            except Exception:
                sp = None
            if sp is not None:
                if sp <= 90:
                    pts += 24; parts.append("SpO₂ ≤90%: +24")
                elif sp <= 95:
                    pts += 8; parts.append("SpO₂ 91–95%: +8")
                else:
                    parts.append("SpO₂ ≥96%: +0")

            if self.recent_inf.get() == "예":
                pts += 17; parts.append("최근 1개월 호흡기 감염: +17")
            else:
                parts.append("최근 1개월 호흡기 감염: +0")

            if self.hb_low.get() == "예":
                pts += 11; parts.append("Hb <10 g/dL: +11")
            else:
                parts.append("Hb ≥10 g/dL: +0")

            inc = self.incision.get()
            if inc == "상복부":
                pts += 15; parts.append("상복부 수술: +15")
            elif inc == "흉부(흉강내)":
                pts += 24; parts.append("흉강내 수술: +24")
            else:
                parts.append("말초 수술: +0")

            try:
                dur = float((self.duration.get() or "").strip())
            except Exception:
                dur = None
            if dur is not None:
                if dur > 180:
                    pts += 23; parts.append("수술시간 >180분: +23")
                elif dur >= 120:
                    pts += 16; parts.append("수술시간 120–180분: +16")
                else:
                    parts.append("수술시간 <120분: +0")

            if self.emerg.get() == "예":
                pts += 8; parts.append("응급수술: +8")
            else:
                parts.append("응급수술 아님: +0")

            if pts < 26:
                risk = "Low (<26)"
            elif pts < 45:
                risk = "Intermediate (26–44)"
            else:
                risk = "High (≥45)"

            self.lbl_ariscat.config(text=f"ARISCAT 점수: {pts}   위험도: {risk}")
            self._set_text(self.txt_ariscat_detail, "세부 점수:\n- " + "\n- ".join(parts))

        # 실시간 계산
        for v in [self.age, self.spo2, self.recent_inf, self.hb_low, self.incision, self.duration, self.emerg]:
            v.trace_add("write", lambda *_: _score())

        btn = ttk.Frame(f)
        btn.grid(row=10, column=0, columnspan=3, sticky="e", padx=10, pady=(0, 10))
        ttk.Button(btn, text="닫기", command=self.destroy).pack(side="right")

        _score()


def open_operability(master):
    try:
        OperabilityWindow(master)
    except Exception as e:
        messagebox.showerror("Operability 오류", f"Operability 창 실행 중 오류가 발생했습니다:\n{e}")





def open_ild_algorithm(master):
    try:
        ILDAlgorithmWindow(master)
    except Exception as e:
        messagebox.showerror("ILD 알고리즘 오류", f"ILD 진단 알고리즘 실행 중 오류가 발생했습니다:\n{e}")

root = tk.Tk()
root.title("호흡기 도우미 By MSW")
root.geometry("1200x750")

# ttk 스타일(탭 폭/여백 포함)
try:
    _style = ttk.Style(root)
    _style.configure("TNotebook.Tab", padding=(18, 8))  # 탭을 넓게
except Exception:
    pass
set_global_font_size(root, CURRENT_FONT_SIZE)



root.columnconfigure(0, weight=1)
root.rowconfigure(0, weight=1)

# ---- 좌우 분할 PanedWindow (왼쪽: 카테고리 / 오른쪽: 규칙+조언) ---- #
main_paned = tk.PanedWindow(root, orient="horizontal")
main_paned.grid(row=0, column=0, sticky="nsew")

# 왼쪽 패널
left_frame = tk.Frame(main_paned, bd=1, relief="solid")
main_paned.add(left_frame, minsize=260)  # 최소 너비 조정 가능

# 오른쪽 패널
right_frame = tk.Frame(main_paned)
main_paned.add(right_frame)

# ----- 왼쪽: 검색 + 카테고리 리스트 ----- #
search_label = tk.Label(left_frame, text="검색", font=("맑은 고딕", 10, "bold"))
search_label.pack(padx=8, pady=(8, 2), anchor="w")

search_var = tk.StringVar()
search_entry = tk.Entry(left_frame, textvariable=search_var)
search_entry.pack(padx=8, fill="x")
search_entry.bind("<Return>", lambda e: on_search())

# 글자 크기 선택(8~20)
font_size_frame = tk.Frame(left_frame)
font_size_frame.pack(padx=8, pady=(6, 2), fill="x")
tk.Label(font_size_frame, text="글자 크기").pack(side="left")
font_size_var = tk.StringVar(value=str(CURRENT_FONT_SIZE))
font_size_combo = ttk.Combobox(
    font_size_frame,
    textvariable=font_size_var,
    values=[str(i) for i in range(8, 21)],
    width=5,
    state="readonly"
)
font_size_combo.pack(side="left", padx=(8, 0))
font_size_combo.bind("<<ComboboxSelected>>", lambda e: set_global_font_size(root, font_size_var.get()))

search_btn_frame = tk.Frame(left_frame)
search_btn_frame.pack(padx=8, pady=(4, 8), fill="x")

search_button = tk.Button(search_btn_frame, text="검색", command=on_search)
search_button.pack(side="left")

reload_button = tk.Button(search_btn_frame, text="엑셀 다시 읽기", command=on_reload_rules)
reload_button.pack(side="right")



lung_tnm_button = tk.Button(search_btn_frame, text="Lung TNM v8", command=lambda: open_lung_tnm(root))
lung_tnm_button.pack(side="right", padx=(8,0))
cat_label = tk.Label(left_frame, text="카테고리", font=("맑은 고딕", 10, "bold"))
cat_label.pack(padx=8, pady=(4, 2), anchor="w")

cat_list_frame = tk.Frame(left_frame)
cat_list_frame.pack(padx=8, pady=(0, 8), fill="both", expand=True)

category_listbox = tk.Listbox(cat_list_frame, exportselection=False)
cat_scroll = tk.Scrollbar(cat_list_frame, command=category_listbox.yview)
category_listbox.config(yscrollcommand=cat_scroll.set)

category_listbox.pack(side="left", fill="both", expand=True)
cat_scroll.pack(side="right", fill="y")

category_listbox.bind("<<ListboxSelect>>", on_category_select)

# ----- 오른쪽 내부: 세로 분할 (위: 규칙 목록, 아래: 조언 내용) ----- #
right_frame.columnconfigure(0, weight=1)
right_frame.rowconfigure(0, weight=1)

right_paned = tk.PanedWindow(right_frame, orient="vertical")
right_paned.grid(row=0, column=0, sticky="nsew")

top_frame = tk.Frame(right_paned)
bottom_frame = tk.Frame(right_paned)

right_paned.add(top_frame, minsize=150)
right_paned.add(bottom_frame, minsize=150)

# ===== 위쪽: 규칙 목록 + 편집 버튼 ===== #
rule_label = tk.Label(top_frame, text="규칙 목록 / 검색 결과", font=("맑은 고딕", 10, "bold"))
rule_label.pack(anchor="w", pady=(0, 4))

rule_list_frame = tk.Frame(top_frame)
rule_list_frame.pack(fill="both", expand=True)

rule_listbox = tk.Listbox(rule_list_frame, height=8, exportselection=False)
rule_scroll = tk.Scrollbar(rule_list_frame, command=rule_listbox.yview)
rule_listbox.config(yscrollcommand=rule_scroll.set)

rule_listbox.pack(side="left", fill="both", expand=True)
rule_scroll.pack(side="right", fill="y")

rule_listbox.bind("<<ListboxSelect>>", on_rule_select)



def open_outpatient_record_writer():
    """외래 기록 작성(템플릿 선택 창) 열기"""
    try:
        OutpatientTemplateChooser(root, rules)
    except Exception as e:
        try:
            messagebox.showerror("오류", f"외래 기록 작성 창을 열 수 없습니다.\n{e}")
        except Exception:
            print("Failed to open outpatient record writer:", e)

btn_frame = tk.Frame(top_frame)
btn_frame.pack(anchor="w", pady=(4, 2))

tk.Button(btn_frame, text="선택 규칙 편집", command=lambda: open_rule_editor("edit")).pack(side="left", padx=4)
tk.Button(btn_frame, text="새 규칙 추가", command=lambda: open_rule_editor("new")).pack(side="left", padx=4)
tk.Button(btn_frame, text="엑셀로 저장", command=on_save_to_excel).pack(side="left", padx=4)

tk.Button(btn_frame, text="외래 기록 작성", command=open_outpatient_record_writer).pack(side="left", padx=4)
tk.Button(btn_frame, text="Lung TNM v8", command=lambda: open_lung_tnm(root)).pack(side="left", padx=4)
tk.Button(btn_frame, text="ILD 진단 알고리즘", command=lambda: open_ild_algorithm(root)).pack(side="left", padx=4)
tk.Button(btn_frame, text="Operability", command=lambda: open_operability(root)).pack(side="left", padx=4)
tk.Button(btn_frame, text="기관지내시경", command=lambda: open_bronchoscopy(root)).pack(side="left", padx=4)
tk.Button(btn_frame, text="임상연구 배정(25.12)", command=lambda: open_clinical_trial_helper(root), bg="#fffACD").pack(side="left", padx=4)

# ===== 아래쪽: 조언 내용 ===== #
advice_label = tk.Label(bottom_frame, text="조언 내용", font=("맑은 고딕", 10, "bold"))
advice_label.pack(anchor="w", pady=(0, 4))

advice_text = scrolledtext.ScrolledText(
    bottom_frame,
    font=("맑은 고딕", 10),
    wrap="word"
)
advice_text.pack(fill="both", expand=True)
advice_text.config(state="disabled")

# ----- 상태 바 ----- #
status_var = tk.StringVar()
status_bar = tk.Label(root, textvariable=status_var, anchor="w", relief="sunken")
status_bar.grid(row=1, column=0, sticky="we")
root.rowconfigure(1, weight=0)

# ================= 초기 데이터 로딩 ================= #
rules = load_rules_from_xlsx()
categories = get_categories(rules) if rules else ["전체"]
displayed_rules = []

refresh_category_list()
status_var.set("준비 완료")

# =============================================================================
# [새 기능] 2025.12 기준 호흡기내과 임상연구 배정 도우미
# =============================================================================
def open_clinical_trial_helper(parent):
    # 팝업창 생성
    win = tk.Toplevel(parent)
    win.title("호흡기센터 임상연구 배정 (2025.12 Ver)")
    win.geometry("1000x800")

    # 스타일 설정 (글자 크기 등)
    style = ttk.Style()
    style.configure("Bold.TLabel", font=("맑은 고딕", 11, "bold"))
    style.configure("Blue.TLabel", foreground="blue", font=("맑은 고딕", 10))
    style.configure("Red.TLabel", foreground="red", font=("맑은 고딕", 10))

    # 탭 컨트롤 생성
    notebook = ttk.Notebook(win)
    notebook.pack(fill="both", expand=True, padx=10, pady=10)

    # ---------------------------------------------------------
    # [탭 1] COPD
    # ---------------------------------------------------------
    tab_copd = ttk.Frame(notebook)
    notebook.add(tab_copd, text="  COPD  ")

    # 1. 레지스트리 (KOCOSS)
    f1 = ttk.LabelFrame(tab_copd, text="1단계: 신규 환자 등록 (KOCOSS)")
    f1.pack(fill="x", padx=10, pady=5)
    
    lbl_kocoss = tk.Label(f1, text="상태를 선택하세요.", justify="left")
    lbl_kocoss.pack(anchor="w", padx=10, pady=5)

    def check_kocoss():
        if var_copd_new.get():
            lbl_kocoss.config(text="✅ [필수] KOCOSS 레지스트리 등록 (담당: 함경은)\n   - 신규 환자 필수 등록\n   - 대상자 중 '노쇠/근감소증 연구' 동시 등록 가능\n   - 유형 분류: TB / BE / Asthma / PRISM / Smoker 중 선택", fg="green", font=("맑은 고딕", 10, "bold"))
        else:
            lbl_kocoss.config(text="기존 등록 환자입니다.", fg="gray", font=("맑은 고딕", 10))

    var_copd_new = tk.BooleanVar()
    tk.Checkbutton(f1, text="기관지확장제 반응 검사 후 FEV1/FVC < 0.7 (신규 진단)", variable=var_copd_new, command=check_kocoss).pack(anchor="w", padx=5)

    # 2. 특수 조건 (박초아 담당)
    f2 = ttk.LabelFrame(tab_copd, text="2단계: 특수 조건 확인 (우선 배정 - 박초아 담당)")
    f2.pack(fill="x", padx=10, pady=5)

    lbl_special = tk.Label(f2, text="해당 사항 없음", justify="left", fg="blue")
    lbl_special.pack(anchor="w", padx=10, pady=5)

    var_home_o2 = tk.BooleanVar()
    var_cough_copd = tk.BooleanVar()
    var_vaccine = tk.BooleanVar()

    def check_copd_special():
        msg = []
        if var_home_o2.get(): msg.append("👉 [가정산소] IIT. 마이숨 (MyBreath) - 재택 모니터링")
        if var_cough_copd.get(): msg.append("👉 [만성기침] IIT. 만성기침 레지스트리")
        if var_vaccine.get(): msg.append("👉 [백신] GSK. Arexvy PMS (50세 이상)")
        
        if msg: lbl_special.config(text="\n".join(msg), fg="red", font=("맑은 고딕", 10, "bold"))
        else: lbl_special.config(text="해당 사항 없음", fg="gray")

    tk.Checkbutton(f2, text="가정 산소 요법 사용 중", variable=var_home_o2, command=check_copd_special).pack(anchor="w", padx=5)
    tk.Checkbutton(f2, text="만성 기침 (8주 이상, 원인미상)", variable=var_cough_copd, command=check_copd_special).pack(anchor="w", padx=5)
    tk.Checkbutton(f2, text="RSV 백신 접종 고려 (50세 이상)", variable=var_vaccine, command=check_copd_special).pack(anchor="w", padx=5)

    # 3. SIT 배정
    f3 = ttk.LabelFrame(tab_copd, text="3단계: 임상시험(SIT) 추가 배정")
    f3.pack(fill="both", expand=True, padx=10, pady=5)

    var_copd_sit = tk.StringVar(value="none")
    lbl_copd_result = tk.Label(f3, text="환자의 임상 상태를 선택해주세요.", justify="left", bg="#f5f5f5", relief="sunken", padx=10, pady=10)
    lbl_copd_result.pack(fill="both", expand=True, padx=5, pady=5)

    def update_copd_sit():
        v = var_copd_sit.get()
        if v == "severe":
            txt = ("🩸 [중증/악화] 생물학적 제제 추천 (담당: 정영진)\n"
                   "   1. AZ_EMBARK (Tezepelumab): 0/4명 (적극 모집 중)\n"
                   "   2. AZ_PRESTO (IRAK4): 0/2명 (신규 기전)\n"
                   "   3. Sanofi Dupilumab / UPSTREAMBIO: 1월 스크리닝 예정")
            lbl_copd_result.config(text=txt, fg="red", font=("맑은 고딕", 11, "bold"))
        elif v == "maint":
            txt = ("💊 [유지 요법] 흡입제 연구 추천 (담당: 최서원)\n"
                   "   1. AZ_THARROS (3제): 12/11명 (추가 TO 확인 필요)\n"
                   "   2. 코오롱 TRIKORE: 30/24명 (마감 임박)\n"
                   "   * 급성 기관지염 동반 시: 대원제약 DWCDS-401 (적극 추천)")
            lbl_copd_result.config(text=txt, fg="blue", font=("맑은 고딕", 11, "bold"))
        elif v == "be":
            txt = ("🧬 [기관지확장증 특화] BI_AIRTIVITY (Cathepsin C) (담당: 정영진)\n   * 12월말~1월초 스크리닝 예정, 적극 모집")
            lbl_copd_result.config(text=txt, fg="#800080", font=("맑은 고딕", 11, "bold"))

    tk.Radiobutton(f3, text="빈번한 급성 악화 (중증/생물학적제제)", variable=var_copd_sit, value="severe", command=update_copd_sit).pack(anchor="w", padx=5, pady=2)
    tk.Radiobutton(f3, text="안정적 유지 치료 필요 (유지/복합제/급성기관지염)", variable=var_copd_sit, value="maint", command=update_copd_sit).pack(anchor="w", padx=5, pady=2)
    tk.Radiobutton(f3, text="기관지확장증 주증상", variable=var_copd_sit, value="be", command=update_copd_sit).pack(anchor="w", padx=5, pady=2)

    # ---------------------------------------------------------
    # [탭 2] 천식 (Asthma)
    # ---------------------------------------------------------
    tab_asthma = ttk.Frame(notebook)
    notebook.add(tab_asthma, text="  천식 (Asthma)  ")

    tk.Label(tab_asthma, text="✅ [기본] TiGER / PRISM / KOSAR (담당: 함경은)\n   * 모든 중증/치료불응성 천식 환자 등록", 
             bg="#e6f3ff", fg="black", font=("맑은 고딕", 11), padx=10, pady=10).pack(fill="x", padx=10, pady=10)

    f_input = ttk.LabelFrame(tab_asthma, text="환자 정보 입력")
    f_input.pack(fill="x", padx=10, pady=10)
    
    f_eos = tk.Frame(f_input)
    f_eos.pack(fill="x", padx=5, pady=5)
    tk.Label(f_eos, text="혈중 호산구(Eosinophil):", font=("맑은 고딕", 10, "bold")).pack(side="left")
    entry_eos = tk.Entry(f_eos, width=10)
    entry_eos.insert(0, "0")
    entry_eos.pack(side="left", padx=5)
    tk.Label(f_eos, text="cells/μL").pack(side="left")

    var_rhinitis = tk.BooleanVar()
    var_cough_asthma = tk.BooleanVar()
    var_uncontrolled = tk.BooleanVar()
    
    tk.Checkbutton(f_input, text="알레르기 비염 동반", variable=var_rhinitis).pack(anchor="w", padx=5)
    tk.Checkbutton(f_input, text="만성 기침 (8주 이상)", variable=var_cough_asthma).pack(anchor="w", padx=5)
    tk.Checkbutton(f_input, text="기존 치료로 조절 안됨 (Uncontrolled)", variable=var_uncontrolled).pack(anchor="w", padx=5)

    lbl_asthma_res = tk.Label(tab_asthma, text="[결과 확인] 버튼을 눌러주세요.", justify="left", bg="#f0f0f0", font=("맑은 고딕", 11), relief="groove", padx=20, pady=20)
    lbl_asthma_res.pack(fill="both", expand=True, padx=10, pady=10)

    def calc_asthma():
        try:
            eos = int(entry_eos.get())
        except:
            eos = 0
        
        res = []
        # 1순위
        if eos >= 300:
            res.append("🌟 [1순위/강력추천] Areteia (EXHALE-2) (담당: 이숙희)\n   - 경구용(먹는 약) 연구, 현재 가장 적극 모집 중")
        
        # 2순위
        if var_rhinitis.get():
            res.append("👃 [비염 동반] 대원제약 DW1807 (담당: 최서원)\n   - 등록 원활 (6/18명)")
        if var_cough_asthma.get():
            res.append("🗣️ [기침] 만성 기침 레지스트리 (담당: 박초아)")
            
        # 3순위
        if var_uncontrolled.get():
            res.append("💉 [생물학적 제제] Sanofi Lunsekimig (LTS17231) (담당: 정영진)\n   - 잔여 슬롯 1명 예상 (Pfizer/Amgen은 대기 필요)")
            
        if not res:
            res.append("👉 특별한 SIT 대상이 아닙니다.\n   1단계 레지스트리(TiGER/PRISM) 등록을 우선 진행하세요.")
            
        lbl_asthma_res.config(text="\n\n".join(res), fg="blue")

    tk.Button(tab_asthma, text="▼ 결과 확인 (Click) ▼", command=calc_asthma, bg="#ffffe0", font=("맑은 고딕", 10, "bold")).pack(fill="x", padx=10, before=lbl_asthma_res)

    # ---------------------------------------------------------
    # [탭 3] 기타 (BE / 기침 / 감기)
    # ---------------------------------------------------------
    tab_etc = ttk.Frame(notebook)
    notebook.add(tab_etc, text="  기타 (BE/기침/감기)  ")

    var_diag = tk.StringVar(value="none")
    f_etc_sel = ttk.LabelFrame(tab_etc, text="주 진단명 선택")
    f_etc_sel.pack(fill="x", padx=10, pady=10)

    lbl_etc_res = tk.Label(tab_etc, text="진단명을 선택하면 배정 가이드가 표시됩니다.", justify="left", font=("맑은 고딕", 11), bg="#f9f9f9", relief="ridge", padx=20, pady=20)
    lbl_etc_res.pack(fill="both", expand=True, padx=10, pady=10)

    def update_etc():
        d = var_diag.get()
        if d == "be":
            lbl_etc_res.config(text="[기관지확장증 (Bronchiectasis)]\n\n1. 객담 녹농균(Pseudomonas) 양성인 경우:\n   👉 AZ_CLERA (담당: 이숙희)\n\n2. 그 외 일반/악화:\n   👉 BI_AIRTIVITY (담당: 정영진) *적극 모집 중", fg="#800080")
        elif d == "cough":
            lbl_etc_res.config(text="[만성 기침 (Chronic Cough)]\n\n👉 만성 기침 레지스트리 (담당: 박초아)\n   * 8주 이상 지속, 원인 미상 또는 난치성 기침\n   * (Bellus 연구는 마감됨)", fg="#a52a2a")
        elif d == "acute":
            lbl_etc_res.config(text="[급성 기관지염 (Acute Bronchitis)]\n\n1. 기저질환 COPD가 있는 경우:\n   👉 대원제약 DWCDS-401 (담당: 최서원)\n\n2. 일반 급성 기관지염:\n   👉 제뉴원사이언스 (담당: 이숙희) *1월 개시 예정", fg="#006400")
        elif d == "ipf":
             lbl_etc_res.config(text="[IPF (특발성 폐섬유증)]\n\n👉 Syndax (SNDX-6352) (담당: 이숙희)\n   * FVC ≥ 50%, 만 40세 이상\n   * 정맥주사(IV), 적극 모집 중", fg="blue")


    tk.Radiobutton(f_etc_sel, text="기관지확장증 (Bronchiectasis)", variable=var_diag, value="be", command=update_etc).pack(anchor="w", padx=5, pady=2)
    tk.Radiobutton(f_etc_sel, text="만성 기침 (Chronic Cough)", variable=var_diag, value="cough", command=update_etc).pack(anchor="w", padx=5, pady=2)
    tk.Radiobutton(f_etc_sel, text="급성 기관지염 (Acute Bronchitis)", variable=var_diag, value="acute", command=update_etc).pack(anchor="w", padx=5, pady=2)
    tk.Radiobutton(f_etc_sel, text="IPF (특발성 폐섬유증)", variable=var_diag, value="ipf", command=update_etc).pack(anchor="w", padx=5, pady=2)

    # 기본 선택 트리거
    check_kocoss()

    
root.mainloop()