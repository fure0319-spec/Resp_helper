import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import os

# 페이지 설정
st.set_page_config(page_title="호흡기내과 임상연구 배정", layout="wide", page_icon="🏥")

# 엑셀 파일 이름
STATUS_EXCEL = "status.xlsx"
CRITERIA_FILE = "criteria.xlsx"

# -----------------------------------------------------------------------------
# 1. 엑셀 데이터 로드 함수 (상단 박스용)
# -----------------------------------------------------------------------------
@st.cache_data(ttl=600)
def load_status_from_excel():
    data = {}
    default_msg = {
        "copd_sit_severe": "데이터 없음", "copd_sit_maint": "데이터 없음",
        "copd_sit_be": "데이터 없음", "asthma_eos": "데이터 없음",
        "asthma_rhinitis": "데이터 없음", "asthma_bio": "데이터 없음",
        "etc_be": "데이터 없음", "etc_cough": "데이터 없음",
        "etc_acute": "데이터 없음", "etc_ipf": "데이터 없음"
    }
    
    if not os.path.exists(STATUS_EXCEL):
        return default_msg

    try:
        wb = load_workbook(STATUS_EXCEL, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=1, values_only=True):
            if row[0] and len(row) > 1:
                key = str(row[0]).strip()
                val = str(row[1]) if row[1] else ""
                # Markdown 줄바꿈 (공백 2칸 + 엔터)
                val = val.replace('\r\n', '\n').replace('\n', '  \n')
                data[key] = val
        wb.close()
    except Exception as e:
        st.error(f"엑셀 읽기 오류: {e}")
        return default_msg
        
    for k, v in default_msg.items():
        if k not in data:
            data[k] = v
    return data

status_data = load_status_from_excel()

# -----------------------------------------------------------------------------
# 2. 웹 화면 구성
# -----------------------------------------------------------------------------

st.title("🏥 건국대병원 호흡기내과 임상연구 배정 도우미")
st.markdown(f"Status Data: `{STATUS_EXCEL}` (2025.12 Ver)")
st.divider()

# 탭 생성
tab1, tab2, tab3 = st.tabs(["🫁 COPD", "🌿 천식 (Asthma)", "🦠 기타 (BE/기침/감기)"])

# [탭 1] COPD
with tab1:
    st.header("COPD 환자 배정")
    col1, col2 = st.columns([1, 1])
    with col1:
        st.subheader("1단계: 레지스트리")
        is_new_copd = st.checkbox("기관지확장제 반응 검사 후 FEV1/FVC < 0.7 (신규 진단)")
        if is_new_copd:
            st.success("✅ **[필수] KOCOSS 레지스트리 등록 (담당: 함경은)**\n\n* 신규 환자 필수 등록\n* 대상자 중 '노쇠/근감소증 연구' 동시 등록 가능")
            st.info("👉 유형 분류: TB / BE / Asthma / PRISM / Smoker 중 선택")
        else:
            st.write("기존 등록 환자입니다.")
    with col2:
        st.subheader("2단계: 특수 조건 (박초아 담당)")
        home_o2 = st.checkbox("가정 산소 요법 사용 중")
        cough_copd = st.checkbox("만성 기침 (8주 이상, 원인미상)")
        vaccine = st.checkbox("RSV 백신 접종 고려 (50세 이상)")
        if home_o2: st.warning("👉 [가정산소] IIT. 마이숨 (MyBreath)")
        if cough_copd: st.warning("👉 [만성기침] IIT. 만성기침 레지스트리")
        if vaccine: st.warning("👉 [백신] GSK. Arexvy PMS")
    
    st.divider()
    st.subheader("3단계: 임상시험(SIT) 추가 배정")
    copd_sit = st.radio("환자의 임상 상태를 선택하세요", 
                        ["선택 안함", "빈번한 급성 악화 (중증/생물학적제제)", "안정적 유지 치료 필요", "기관지확장증 주증상"])
    if copd_sit == "빈번한 급성 악화 (중증/생물학적제제)": st.error(status_data["copd_sit_severe"])
    elif copd_sit == "안정적 유지 치료 필요": st.info(status_data["copd_sit_maint"])
    elif copd_sit == "기관지확장증 주증상": st.success(status_data["copd_sit_be"])

# [탭 2] 천식
with tab2:
    st.header("천식 (Asthma) 환자 배정")
    st.info("✅ **[기본] TiGER / PRISM / KOSAR (담당: 함경은)**\n\n* 모든 중증/치료불응성 천식 환자 등록")
    st.markdown("### 환자 정보 입력")
    col_a, col_b = st.columns([1, 2])
    with col_a: eos_input = st.number_input("혈중 호산구(Eosinophil)", min_value=0, step=10)
    with col_b:
        has_rhinitis = st.checkbox("알레르기 비염 동반")
        has_cough_asthma = st.checkbox("만성 기침 (8주 이상) 동반")
        is_uncontrolled = st.checkbox("기존 치료로 조절 안됨 (Uncontrolled)")
    st.markdown("### 배정 결과")
    results = []
    if eos_input >= 300: st.success(status_data["asthma_eos"]); results.append(True)
    if has_rhinitis: st.warning(status_data["asthma_rhinitis"]); results.append(True)
    if has_cough_asthma: st.warning(status_data["etc_cough"]); results.append(True)
    if is_uncontrolled: st.error(status_data["asthma_bio"]); results.append(True)
    if not results: st.info("👉 특별한 SIT 대상이 아닙니다. 1단계 레지스트리 등록을 우선 진행하세요.")

# [탭 3] 기타
with tab3:
    st.header("기타 (BE / 기침 / 급성기관지염 / IPF)")
    diagnosis = st.radio("주 진단명을 선택하세요", 
                         ["기관지확장증 (Bronchiectasis)", "만성 기침 (Chronic Cough)", "급성 기관지염 (Acute Bronchitis)", "IPF (특발성 폐섬유증)"])
    st.markdown("### 배정 가이드")
    if diagnosis == "기관지확장증 (Bronchiectasis)": st.success(status_data["etc_be"])
    elif diagnosis == "만성 기침 (Chronic Cough)": st.warning(status_data["etc_cough"])
    elif diagnosis == "급성 기관지염 (Acute Bronchitis)": st.info(status_data["etc_acute"])
    elif diagnosis == "IPF (특발성 폐섬유증)": st.error(status_data["etc_ipf"])

# ==========================================
# [통합 기능] 하단 상세 엑셀 파일 표시 (HTML Table 방식 적용)
# ==========================================
st.divider()
st.header("📑 연구별 상세 선정/제외 기준 (Reference)")

if os.path.exists(CRITERIA_FILE):
    try:
        target_sheets = ["천식", "COPD", "BE기침기관지염", "기타(IPF, 암)", "예정"]
        all_dfs = []
        
        for sheet in target_sheets:
            try:
                # 1. 엑셀 읽기 (모든 열 읽음)
                raw_df = pd.read_excel(CRITERIA_FILE, sheet_name=sheet).astype(str)
                raw_df = raw_df.replace("nan", "")
                
                # 2. 컬럼 매핑 (자동 찾기 + 강제 매핑)
                # 교수님 요청: B열=선정기준, C열=제외기준
                # 실제 엑셀 구조: [질환(A), 연구제목(B), 선정기준(C), 제외기준(D), 비고(E)...] 형태일 가능성이 높음
                # 따라서 B, C가 아니라 "연구제목", "선정기준", "제외기준" 열을 찾아서 가져옴
                
                # 데이터프레임 컬럼 이름 정리 (공백 제거)
                raw_df.columns = [c.strip() for c in raw_df.columns]
                
                # 필요한 컬럼만 추출하기 위한 로직
                new_row = {}
                new_row['분류'] = sheet
                
                # (1) 연구 과제명 (보통 B열 또는 '과제', '제목' 포함된 열)
                # B열(index 1)을 기본으로 잡되, 헤더 이름 확인
                if len(raw_df.columns) > 1:
                     new_row['연구 과제'] = raw_df.iloc[:, 1] # B열
                else:
                     new_row['연구 과제'] = ""

                # (2) 선정 기준 (교수님이 B열이라고 하셨지만, 일반적으론 C열. 
                #     혹시 모르니 '선정' 글자가 있는 열을 우선 찾고, 없으면 C열 사용)
                col_selection = [c for c in raw_df.columns if "선정" in c]
                if col_selection:
                    new_row['선정 기준'] = raw_df[col_selection[0]]
                elif len(raw_df.columns) > 2:
                    new_row['선정 기준'] = raw_df.iloc[:, 2] # C열
                else:
                    new_row['선정 기준'] = ""

                # (3) 제외 기준 ('제외' 글자 우선, 없으면 D열)
                col_exclusion = [c for c in raw_df.columns if "제외" in c]
                if col_exclusion:
                    new_row['제외 기준'] = raw_df[col_exclusion[0]]
                elif len(raw_df.columns) > 3:
                    new_row['제외 기준'] = raw_df.iloc[:, 3] # D열
                else:
                    new_row['제외 기준'] = ""

                # 재조립
                temp_df = pd.DataFrame(new_row)
                all_dfs.append(temp_df)
                
            except ValueError: continue
        
        if all_dfs:
            df = pd.concat(all_dfs, ignore_index=True)
            
            # 검색 기능
            search_query = st.text_input("🔍 키워드 검색", placeholder="예: 천식, COPD, 호산구")
            
            if search_query:
                query = search_query.strip()
                mask = df.apply(lambda row: row.astype(str).str.contains(query, case=False).any(), axis=1)
                df_display = df[mask]
            else:
                df_display = df

            st.caption(f"총 **{len(df_display)}**건의 연구 기준")

            # [핵심 수정] HTML Table로 변환하여 렌더링 (자동 줄바꿈, 높이 자동 조절)
            # 스타일 정의: Arial, 10pt, 서식 없음(흰배경), 테두리
            table_html = df_display.to_html(index=False, escape=False)
            
            # CSS 적용 (테이블 폭 100%, 폰트 설정, 줄바꿈 허용)
            custom_css = """
            <style>
                table {
                    width: 100%;
                    border-collapse: collapse;
                    font-family: Arial, sans-serif;
                    font-size: 10pt;
                    color: black;
                    background-color: white;
                }
                th {
                    background-color: #f0f2f6;
                    border: 1px solid #ddd;
                    padding: 8px;
                    text-align: left;
                    font-weight: bold;
                }
                td {
                    border: 1px solid #ddd;
                    padding: 8px;
                    vertical-align: top; /* 내용이 길면 위쪽 정렬 */
                    white-space: pre-wrap; /* 줄바꿈 허용 (가장 중요) */
                    word-wrap: break-word;
                }
                tr:nth-child(even) {background-color: #ffffff;}
            </style>
            """
            
            st.markdown(custom_css + table_html, unsafe_allow_html=True)

        else:
            st.warning("⚠️ 지정된 시트(탭)를 찾을 수 없습니다.")

    except Exception as e:
        st.error(f"오류 발생: {e}")
else:
    st.info("ℹ️ 상세 기준 파일(criteria.xlsx)이 없습니다.")